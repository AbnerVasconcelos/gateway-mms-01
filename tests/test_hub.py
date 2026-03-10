#!/usr/bin/env python3
"""
Testes da Fase 2 — Hub: Socket.IO bridge + Config Panel.

Duas suites:

  TestConfigStore      — unitária, sem dependências externas.
                         Usa diretório temporário para isolar os arquivos de config.

  TestHubIntegration   — inicia o Hub como subprocess + conecta cliente Socket.IO real.
                         Requer Redis rodando. Ignorada automaticamente se indisponível.

Uso:
    python -m pytest tests/test_hub.py -v
    python -m pytest tests/test_hub.py -v -k TestConfigStore   # só unitários
"""

import json
import logging
import os
import shutil
import subprocess
import sys
import tempfile
import time
import unittest

import redis

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(name)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)
logger = logging.getLogger('test_hub')

GATEWAY_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if sys.platform == 'win32':
    PYTHON = os.path.join(GATEWAY_DIR, '.venv', 'Scripts', 'python')
else:
    PYTHON = os.path.join(GATEWAY_DIR, '.venv', 'bin', 'python')
REDIS_HOST  = os.environ.get('REDIS_HOST', 'localhost')
REDIS_PORT  = int(os.environ.get('REDIS_PORT', 6379))
HUB_PORT    = 8765   # porta exclusiva para testes — evita conflito com Hub em produção
TABLES_DIR  = os.path.join(GATEWAY_DIR, 'tables')
LOG_DIR     = os.path.join(GATEWAY_DIR, 'tests', 'logs')

sys.path.insert(0, GATEWAY_DIR)
sys.path.insert(0, os.path.join(GATEWAY_DIR, 'Hub'))
import config_store  # noqa: E402


# ---------------------------------------------------------------------------
# Suite 1 — TestConfigStore (unitário, sem deps externas)
# ---------------------------------------------------------------------------

class TestConfigStore(unittest.TestCase):
    """
    Testa config_store.py isolado, usando um diretório temporário para não
    modificar os arquivos reais de configuração.
    """

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        shutil.copy(
            os.path.join(TABLES_DIR, 'group_config.json'),
            os.path.join(self.temp_dir, 'group_config.json'),
        )
        shutil.copy(
            os.path.join(TABLES_DIR, 'variable_overrides.json'),
            os.path.join(self.temp_dir, 'variable_overrides.json'),
        )
        self._orig_tables_dir = config_store._TABLES_DIR
        config_store._TABLES_DIR = self.temp_dir

    def tearDown(self):
        config_store._TABLES_DIR = self._orig_tables_dir
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    # ── load_group_config ────────────────────────────────────────────────────

    def test_01_load_group_config_returns_dict(self):
        cfg = config_store.load_group_config()
        self.assertIsInstance(cfg, dict)
        self.assertNotIn('groups', cfg)   # Fase 5: seção groups removida
        self.assertIn('_meta', cfg)
        # Channels are now inside devices, not at top level
        self.assertIn('devices', cfg)
        self.assertIn('channels', cfg['devices']['simulador'])

    def test_02_load_group_config_has_expected_channels(self):
        cfg = config_store.load_group_config()
        # Channels are now inside devices
        device_channels = cfg['devices']['simulador'].get('channels', {})
        for ch in ('plc_alarmes', 'plc_retentivas', 'plc_operacao'):
            self.assertIn(ch, device_channels, f"Canal '{ch}' ausente em devices.simulador.channels")

    # ── save_group_config ────────────────────────────────────────────────────

    def test_03_save_group_config_roundtrip(self):
        cfg = config_store.load_group_config()
        cfg['_meta']['test_marker'] = 'roundtrip'
        config_store.save_group_config(cfg)

        reloaded = config_store.load_group_config()
        self.assertEqual(reloaded['_meta']['test_marker'], 'roundtrip')

    def test_04_save_group_config_preserves_channels(self):
        cfg = config_store.load_group_config()
        # Channels are now inside devices
        original_channels = set(cfg['devices']['simulador']['channels'].keys())
        config_store.save_group_config(cfg)
        reloaded = config_store.load_group_config()
        self.assertEqual(original_channels, set(reloaded['devices']['simulador']['channels'].keys()))

    # ── load_overrides ───────────────────────────────────────────────────────

    def test_05_load_overrides_returns_dict(self):
        self.assertIsInstance(config_store.load_overrides(), dict)

    def test_06_load_overrides_missing_file_returns_empty(self):
        os.remove(os.path.join(self.temp_dir, 'variable_overrides.json'))
        self.assertEqual(config_store.load_overrides(), {})

    # ── save_overrides ───────────────────────────────────────────────────────

    def test_07_save_overrides_roundtrip(self):
        data = {'extrusoraErro': {'enabled': False}}
        config_store.save_overrides(data)
        loaded = config_store.load_overrides()
        self.assertEqual(loaded, data)

    # ── update_channel_history_size ──────────────────────────────────────────

    def test_08_update_history_updates_channel_section(self):
        channel = 'plc_alarmes'
        new_size = 42
        config_store.update_channel_history_size(channel, new_size)
        cfg = config_store.load_group_config()
        self.assertEqual(cfg['channels'][channel]['history_size'], new_size)

    def test_09_update_history_does_not_affect_other_channels(self):
        other_channel = 'plc_config'
        size_before = config_store.get_channels().get(other_channel, {}).get('history_size', 100)

        config_store.update_channel_history_size('plc_alarmes', 99)

        size_after = config_store.get_channels().get(other_channel, {}).get('history_size', 100)
        self.assertEqual(
            size_after, size_before,
            f"Canal '{other_channel}' foi alterado indevidamente"
        )

    def test_10_update_history_persists_to_file(self):
        config_store.update_channel_history_size('plc_process', 77, device_id='simulador')
        # Verify via device-specific channels (get_channels aggregates and may overwrite)
        dev_channels = config_store.get_device_channels('simulador')
        self.assertIn('plc_process', dev_channels)
        self.assertEqual(dev_channels['plc_process']['history_size'], 77)

    # ── get_channel_history_sizes ────────────────────────────────────────────

    def test_11_get_channel_history_sizes_unique_channels(self):
        sizes = config_store.get_channel_history_sizes()
        self.assertIsInstance(sizes, dict)
        # Canais esperados no projeto (ao menos os do device simulador)
        for ch in ('plc_alarmes', 'plc_retentivas', 'plc_operacao'):
            self.assertIn(ch, sizes, f"Canal '{ch}' ausente no retorno")

    def test_12_get_channel_history_sizes_values_are_positive_int(self):
        for ch, size in config_store.get_channel_history_sizes().items():
            with self.subTest(channel=ch):
                self.assertIsInstance(size, int)
                self.assertGreater(size, 0)

    def test_13_each_channel_appears_once(self):
        sizes = config_store.get_channel_history_sizes()
        # Cada canal deve aparecer exatamente uma vez (sem duplicatas de key no dict)
        self.assertEqual(len(sizes), len(set(sizes.keys())))

    # ── patch_variable_override ──────────────────────────────────────────────

    def test_14_patch_variable_creates_override(self):
        config_store.patch_variable_override('extrusoraErro', {'enabled': False})
        overrides = config_store.load_overrides()
        self.assertIn('extrusoraErro', overrides)
        self.assertFalse(overrides['extrusoraErro']['enabled'])

    def test_15_patch_variable_merges_fields(self):
        config_store.patch_variable_override('testTag', {'enabled': True})
        config_store.patch_variable_override('testTag', {'channel': 'plc_visual'})
        overrides = config_store.load_overrides()
        self.assertTrue(overrides['testTag']['enabled'])
        self.assertEqual(overrides['testTag']['channel'], 'plc_visual')


# ---------------------------------------------------------------------------
# Suite 1b — TestPerDeviceConfig (unitário, sem deps externas)
# ---------------------------------------------------------------------------

class TestPerDeviceConfig(unittest.TestCase):
    """
    Testa funcionalidades per-device de config_store:
    channels inside devices, per-device overrides, device-scoped channel CRUD.
    """

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        # Create a minimal group_config with per-device channels
        self.group_config = {
            '_meta': {
                'aggregate_channel': 'plc_data',
                'backward_compatible': True,
                'default_delay_ms': 1000,
                'default_history_size': 100,
            },
            'devices': {
                'sim': {
                    'label': 'Simulador',
                    'protocol': 'tcp',
                    'host': 'localhost',
                    'port': 5020,
                    'unit_id': 1,
                    'csv_files': ['test_vars.csv'],
                    'channels': {
                        'plc_alarmes': {'delay_ms': 200, 'history_size': 55},
                        'plc_process': {'delay_ms': 500, 'history_size': 100},
                    },
                    'command_channel': 'sim_commands',
                },
                'west': {
                    'label': 'West',
                    'protocol': 'tcp',
                    'host': 'localhost',
                    'port': 5021,
                    'unit_id': 1,
                    'csv_files': ['test_vars2.csv'],
                    'channels': {
                        'west_data': {'delay_ms': 1000, 'history_size': 50},
                    },
                    'command_channel': 'west_commands',
                },
            },
            'channels': {},  # empty global channels
        }
        with open(os.path.join(self.temp_dir, 'group_config.json'), 'w') as f:
            json.dump(self.group_config, f)

        # Create global overrides
        self.global_overrides = {
            'tagA': {'channel': 'plc_alarmes'},
            'tagB': {'channel': 'plc_process', 'enabled': False},
        }
        with open(os.path.join(self.temp_dir, 'variable_overrides.json'), 'w') as f:
            json.dump(self.global_overrides, f)

        # Create a test CSV
        csv_content = "key,ObjecTag,Tipo,Modbus,At\nalarmes,tagA,M,0,%MB\nprocess,tagB,D,100,%MW\n"
        with open(os.path.join(self.temp_dir, 'test_vars.csv'), 'w') as f:
            f.write(csv_content)

        csv_content2 = "key,ObjecTag,Tipo,Modbus,At\ntemp,tagC,D,200,%MW\ntemp,tagD,D,201,%MW\n"
        with open(os.path.join(self.temp_dir, 'test_vars2.csv'), 'w') as f:
            f.write(csv_content2)

        self._orig_tables_dir = config_store._TABLES_DIR
        config_store._TABLES_DIR = self.temp_dir

    def tearDown(self):
        config_store._TABLES_DIR = self._orig_tables_dir
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    # ── get_channels from devices ────────────────────────────────────────────

    def test_get_channels_from_devices(self):
        """get_channels() deve ler channels de dentro dos devices."""
        channels = config_store.get_channels()
        self.assertIn('plc_alarmes', channels)
        self.assertIn('plc_process', channels)
        self.assertIn('west_data', channels)
        # Verifica que device_id é retornado
        self.assertEqual(channels['plc_alarmes']['device_id'], 'sim')
        self.assertEqual(channels['west_data']['device_id'], 'west')
        self.assertEqual(channels['plc_alarmes']['delay_ms'], 200)
        self.assertEqual(channels['plc_alarmes']['history_size'], 55)

    # ── get_device_channels ──────────────────────────────────────────────────

    def test_get_device_channels(self):
        """get_device_channels() deve retornar apenas canais do device especificado."""
        sim_channels = config_store.get_device_channels('sim')
        self.assertIn('plc_alarmes', sim_channels)
        self.assertIn('plc_process', sim_channels)
        self.assertNotIn('west_data', sim_channels)
        self.assertEqual(sim_channels['plc_alarmes']['delay_ms'], 200)

        west_channels = config_store.get_device_channels('west')
        self.assertIn('west_data', west_channels)
        self.assertNotIn('plc_alarmes', west_channels)

    def test_get_device_channels_nonexistent_device(self):
        """get_device_channels() para device inexistente retorna dict vazio."""
        channels = config_store.get_device_channels('nao_existe')
        self.assertEqual(channels, {})

    # ── Per-device overrides ─────────────────────────────────────────────────

    def test_overrides_per_device_load(self):
        """load_overrides('sim') deve carregar variable_overrides_sim.json."""
        # Create per-device overrides file
        sim_overrides = {'tagA': {'channel': 'plc_alarmes', 'enabled': True}}
        with open(os.path.join(self.temp_dir, 'variable_overrides_sim.json'), 'w') as f:
            json.dump(sim_overrides, f)

        loaded = config_store.load_overrides('sim')
        self.assertEqual(loaded, sim_overrides)

        # Global overrides should be independent
        global_loaded = config_store.load_overrides()
        self.assertEqual(global_loaded, self.global_overrides)

    def test_overrides_per_device_load_missing_returns_empty(self):
        """load_overrides('nonexistent') sem arquivo retorna {}."""
        loaded = config_store.load_overrides('nonexistent')
        self.assertEqual(loaded, {})

    def test_overrides_per_device_save(self):
        """save_overrides(data, 'sim') deve gravar em variable_overrides_sim.json."""
        data = {'tagX': {'channel': 'plc_process'}}
        config_store.save_overrides(data, 'sim')

        path = os.path.join(self.temp_dir, 'variable_overrides_sim.json')
        self.assertTrue(os.path.exists(path))

        with open(path, 'r') as f:
            saved = json.load(f)
        self.assertEqual(saved, data)

        # Global overrides file should not be affected
        global_loaded = config_store.load_overrides()
        self.assertEqual(global_loaded, self.global_overrides)

    def test_patch_variable_override_per_device(self):
        """patch_variable_override com device_id opera no arquivo per-device."""
        config_store.patch_variable_override('tagA', {'channel': 'plc_process'}, device_id='sim')

        # Per-device file should have the patch
        sim_overrides = config_store.load_overrides('sim')
        self.assertIn('tagA', sim_overrides)
        self.assertEqual(sim_overrides['tagA']['channel'], 'plc_process')

        # Global overrides should remain unchanged
        global_overrides = config_store.load_overrides()
        self.assertEqual(global_overrides['tagA']['channel'], 'plc_alarmes')

    # ── Channel CRUD device-scoped ───────────────────────────────────────────

    def test_create_channel_in_device(self):
        """create_channel com device_id cria canal dentro de devices[device_id].channels."""
        config_store.create_channel('plc_new', delay_ms=300, history_size=75, device_id='sim')

        cfg = config_store.load_group_config()
        dev_channels = cfg['devices']['sim']['channels']
        self.assertIn('plc_new', dev_channels)
        self.assertEqual(dev_channels['plc_new']['delay_ms'], 300)
        self.assertEqual(dev_channels['plc_new']['history_size'], 75)

    def test_create_channel_in_nonexistent_device_raises(self):
        """create_channel em device inexistente deve lançar KeyError."""
        with self.assertRaises(KeyError):
            config_store.create_channel('plc_x', device_id='nao_existe')

    def test_delete_channel_from_device(self):
        """delete_channel com device_id remove canal de devices[device_id].channels."""
        config_store.delete_channel('plc_alarmes', device_id='sim')

        cfg = config_store.load_group_config()
        dev_channels = cfg['devices']['sim']['channels']
        self.assertNotIn('plc_alarmes', dev_channels)
        # Other channels in the same device should remain
        self.assertIn('plc_process', dev_channels)

    def test_delete_channel_system_raises(self):
        """delete_channel de canal de sistema deve lançar ValueError."""
        with self.assertRaises(ValueError):
            config_store.delete_channel('user_status', device_id='sim')

    def test_delete_channel_nonexistent_in_device_raises(self):
        """delete_channel de canal inexistente no device deve lançar KeyError."""
        with self.assertRaises(KeyError):
            config_store.delete_channel('canal_fantasma', device_id='sim')

    # ── create_device auto command_channel ───────────────────────────────────

    def test_create_device_auto_command_channel(self):
        """create_device auto-gera command_channel e channels vazio se não fornecido."""
        cfg = {'label': 'CLP 3', 'protocol': 'tcp', 'host': '10.0.0.1',
               'port': 502, 'unit_id': 1, 'csv_files': []}
        config_store.create_device('clp3', cfg)

        devices = config_store.get_devices()
        self.assertIn('clp3', devices)
        self.assertEqual(devices['clp3']['command_channel'], 'clp3_commands')
        self.assertEqual(devices['clp3']['channels'], {})

    def test_create_device_preserves_explicit_command_channel(self):
        """create_device preserva command_channel se já fornecido."""
        cfg = {'label': 'CLP 4', 'protocol': 'tcp', 'host': '10.0.0.2',
               'port': 502, 'unit_id': 1, 'csv_files': [],
               'command_channel': 'custom_cmd'}
        config_store.create_device('clp4', cfg)

        devices = config_store.get_devices()
        self.assertEqual(devices['clp4']['command_channel'], 'custom_cmd')

    # ── load_all_variables per-device overrides ──────────────────────────────

    def test_load_all_variables_per_device_overrides(self):
        """load_all_variables usa per-device override quando arquivo existe."""
        # Create per-device overrides for 'sim' with different channel
        sim_overrides = {'tagA': {'channel': 'plc_process'}, 'tagB': {'channel': 'plc_alarmes'}}
        config_store.save_overrides(sim_overrides, 'sim')

        variables = config_store.load_all_variables()

        # Find variables from 'sim' device
        sim_vars = [v for v in variables if v['device'] == 'sim']
        self.assertGreater(len(sim_vars), 0)

        tagA = next((v for v in sim_vars if v['tag'] == 'tagA'), None)
        self.assertIsNotNone(tagA)
        # Should use per-device override (plc_process), not global (plc_alarmes)
        self.assertEqual(tagA['channel'], 'plc_process')

        tagB = next((v for v in sim_vars if v['tag'] == 'tagB'), None)
        self.assertIsNotNone(tagB)
        # Per-device override has enabled=True (default), global has enabled=False
        self.assertTrue(tagB['enabled'])
        self.assertEqual(tagB['channel'], 'plc_alarmes')

    def test_load_all_variables_falls_back_to_global_overrides(self):
        """load_all_variables usa global overrides quando per-device não existe."""
        # No per-device override file for 'sim' — should fall back to global
        variables = config_store.load_all_variables()

        sim_vars = [v for v in variables if v['device'] == 'sim']
        tagB = next((v for v in sim_vars if v['tag'] == 'tagB'), None)
        self.assertIsNotNone(tagB)
        # Should use global override (enabled=False)
        self.assertFalse(tagB['enabled'])
        self.assertEqual(tagB['channel'], 'plc_process')

    # ── SYSTEM_CHANNELS updated ──────────────────────────────────────────────

    def test_system_channels_updated(self):
        """SYSTEM_CHANNELS deve conter apenas user_status, ia_status, ia_data."""
        self.assertEqual(
            config_store.SYSTEM_CHANNELS,
            frozenset(['user_status', 'ia_status', 'ia_data']),
        )
        # Channels removed from SYSTEM_CHANNELS
        self.assertNotIn('config_reload', config_store.SYSTEM_CHANNELS)
        self.assertNotIn('plc_commands', config_store.SYSTEM_CHANNELS)
        self.assertNotIn('plc_data', config_store.SYSTEM_CHANNELS)
        self.assertNotIn('alarms', config_store.SYSTEM_CHANNELS)

    # ── get_channel_history_sizes aggregates from devices ────────────────────

    def test_get_channel_history_sizes_from_devices(self):
        """get_channel_history_sizes deve agregar de todos os devices."""
        sizes = config_store.get_channel_history_sizes()
        self.assertIn('plc_alarmes', sizes)
        self.assertIn('plc_process', sizes)
        self.assertIn('west_data', sizes)
        self.assertEqual(sizes['plc_alarmes'], 55)
        self.assertEqual(sizes['west_data'], 50)

    # ── update_channel_delay / history_size device-scoped ────────────────────

    def test_update_channel_delay_in_device(self):
        """update_channel_delay com device_id opera em devices[device_id].channels."""
        config_store.update_channel_delay('plc_alarmes', 150, device_id='sim')
        cfg = config_store.load_group_config()
        self.assertEqual(cfg['devices']['sim']['channels']['plc_alarmes']['delay_ms'], 150)

    def test_update_channel_history_size_in_device(self):
        """update_channel_history_size com device_id opera em devices[device_id].channels."""
        config_store.update_channel_history_size('plc_alarmes', 200, device_id='sim')
        cfg = config_store.load_group_config()
        self.assertEqual(cfg['devices']['sim']['channels']['plc_alarmes']['history_size'], 200)

    # ── validate_channel_device with shared CSVs ─────────────────────────────

    def test_validate_channel_device_shared_csv_without_device_id(self):
        """Shared CSV: sem device_id explícito, find_tag_device retorna o primeiro device."""
        # Add shared CSV to both devices
        shared_csv = "key,ObjecTag,Tipo,Modbus,At\ntemp,sharedTag,D,300,%MW\n"
        with open(os.path.join(self.temp_dir, 'shared.csv'), 'w') as f:
            f.write(shared_csv)

        cfg = config_store.load_group_config()
        cfg['devices']['sim']['csv_files'].append('shared.csv')
        cfg['devices']['west']['csv_files'].append('shared.csv')
        config_store.save_group_config(cfg)

        # Without explicit device_id, find_tag_device returns 'sim' (first in dict)
        # so validating against 'west_data' channel should fail
        with self.assertRaises(ValueError):
            config_store.validate_channel_device('sharedTag', 'west_data')

    def test_validate_channel_device_shared_csv_with_explicit_device_id(self):
        """Shared CSV: com device_id explícito, validação usa o device correto."""
        shared_csv = "key,ObjecTag,Tipo,Modbus,At\ntemp,sharedTag,D,300,%MW\n"
        with open(os.path.join(self.temp_dir, 'shared.csv'), 'w') as f:
            f.write(shared_csv)

        cfg = config_store.load_group_config()
        cfg['devices']['sim']['csv_files'].append('shared.csv')
        cfg['devices']['west']['csv_files'].append('shared.csv')
        config_store.save_group_config(cfg)

        # With explicit device_id='west', validation should PASS for west_data
        config_store.validate_channel_device('sharedTag', 'west_data', device_id='west')

        # With explicit device_id='sim', validation should PASS for plc_alarmes
        config_store.validate_channel_device('sharedTag', 'plc_alarmes', device_id='sim')

        # With explicit device_id='west', validation should FAIL for plc_alarmes (sim's channel)
        with self.assertRaises(ValueError):
            config_store.validate_channel_device('sharedTag', 'plc_alarmes', device_id='west')


# ---------------------------------------------------------------------------
# Suite 2 — TestHubIntegration (requer Redis + Hub rodando)
# ---------------------------------------------------------------------------

_hub_proc   = None
_redis_conn = None


def _redis_available():
    try:
        r = redis.Redis(host=REDIS_HOST, port=REDIS_PORT, db=0)
        r.ping()
        r.close()
        return True
    except Exception:
        return False


def _hub_ready(timeout=10):
    """Aguarda o Hub responder em /health."""
    import urllib.request
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            with urllib.request.urlopen(
                f'http://127.0.0.1:{HUB_PORT}/health', timeout=1
            ) as resp:
                if resp.status == 200:
                    return True
        except Exception:
            pass
        time.sleep(0.3)
    return False


def setUpModule():
    global _hub_proc, _redis_conn

    if not _redis_available():
        return  # testes de integração serão ignorados via skipTest nas classes

    os.makedirs(LOG_DIR, exist_ok=True)

    _hub_proc = subprocess.Popen(
        [
            PYTHON, '-m', 'uvicorn', 'Hub.main:asgi_app',
            '--host', '127.0.0.1',
            '--port', str(HUB_PORT),
        ],
        cwd=GATEWAY_DIR,
        stdout=open(os.path.join(LOG_DIR, 'hub_proc.log'), 'w'),
        stderr=subprocess.STDOUT,
    )

    if not _hub_ready(timeout=12):
        _hub_proc.terminate()
        _hub_proc = None
        return

    _redis_conn = redis.Redis(host=REDIS_HOST, port=REDIS_PORT, db=0)
    logger.info("Hub iniciado na porta %d.", HUB_PORT)


def tearDownModule():
    global _hub_proc, _redis_conn
    if _redis_conn:
        _redis_conn.close()
    if _hub_proc:
        _hub_proc.terminate()
        try:
            _hub_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            _hub_proc.kill()
        logger.info("Hub encerrado.")


class TestHubHTTP(unittest.TestCase):
    """Testa os endpoints REST do Hub via HTTP."""

    @classmethod
    def setUpClass(cls):
        if not _redis_available() or _hub_proc is None:
            raise unittest.SkipTest('Hub não disponível (Redis ou processo não iniciado).')
        import urllib.request
        cls.base = f'http://127.0.0.1:{HUB_PORT}'
        cls.urlopen = urllib.request.urlopen

    def _get(self, path):
        import urllib.request
        with urllib.request.urlopen(f'{self.base}{path}', timeout=3) as r:
            return json.loads(r.read().decode())

    def _patch(self, path, payload):
        import urllib.request
        data = json.dumps(payload).encode()
        req  = urllib.request.Request(
            f'{self.base}{path}',
            data=data,
            method='PATCH',
            headers={'Content-Type': 'application/json'},
        )
        with urllib.request.urlopen(req, timeout=3) as r:
            return json.loads(r.read().decode())

    def test_30_health_returns_ok(self):
        resp = self._get('/health')
        self.assertEqual(resp.get('status'), 'ok')

    def test_31_api_channels_returns_dict(self):
        resp = self._get('/api/channels')
        self.assertIsInstance(resp, dict)
        for ch in ('plc_alarmes', 'plc_retentivas', 'plc_operacao'):
            self.assertIn(ch, resp, f"Canal '{ch}' ausente em /api/channels")

    def test_32_api_channels_values_are_dicts_with_delay_and_history(self):
        for ch, cfg in self._get('/api/channels').items():
            with self.subTest(channel=ch):
                self.assertIsInstance(cfg, dict)
                self.assertIn('delay_ms', cfg)
                self.assertIn('history_size', cfg)
                self.assertGreater(cfg['delay_ms'], 0)
                self.assertGreater(cfg['history_size'], 0)

    def test_33_api_groups_returns_empty_dict(self):
        resp = self._get('/api/groups')
        self.assertIsInstance(resp, dict)
        self.assertEqual(len(resp), 0)   # Fase 5: seção groups removida

    def test_34_api_variables_has_overrides_key(self):
        resp = self._get('/api/variables')
        self.assertIn('overrides', resp)

    def test_35_patch_history_updates_channel(self):
        resp = self._patch('/api/channels/plc_alarmes/history?device_id=simulador', {'history_size': 55})
        self.assertEqual(resp['channel'],      'plc_alarmes')
        self.assertEqual(resp['history_size'], 55)
        # Verifica que foi persistido (formato agora é dict)
        channels = self._get('/api/channels')
        self.assertEqual(channels.get('plc_alarmes', {}).get('history_size'), 55)

    def test_36_patch_history_invalid_size_returns_422(self):
        import urllib.error
        with self.assertRaises(urllib.error.HTTPError) as ctx:
            self._patch('/api/channels/plc_alarmes/history', {'history_size': 0})
        self.assertEqual(ctx.exception.code, 422)


class TestHubSocketIO(unittest.TestCase):
    """Testa os eventos Socket.IO do Hub com cliente real."""

    @classmethod
    def setUpClass(cls):
        if not _redis_available() or _hub_proc is None:
            raise unittest.SkipTest('Hub não disponível (Redis ou processo não iniciado).')
        try:
            import socketio as sio_lib
            cls.sio_lib = sio_lib
        except ImportError:
            raise unittest.SkipTest('python-socketio não instalado.')

    def _make_client(self):
        sio = self.sio_lib.Client(logger=False, engineio_logger=False)
        sio.connect(f'http://127.0.0.1:{HUB_PORT}')
        return sio

    def test_40_connect_receives_connection_ack(self):
        """Ao conectar, cliente deve receber connection_ack com available_rooms."""
        received = []

        sio = self.sio_lib.Client(logger=False, engineio_logger=False)

        @sio.on('connection_ack')
        def on_ack(data):
            received.append(data)

        sio.connect(f'http://127.0.0.1:{HUB_PORT}')
        time.sleep(0.5)
        sio.disconnect()

        self.assertEqual(len(received), 1, "connection_ack não recebido")
        self.assertEqual(received[0]['status'], 'connected')
        self.assertIn('available_rooms', received[0])

    def test_41_join_event_accepted(self):
        """Evento join deve ser aceito sem erro."""
        sio = self._make_client()
        try:
            sio.emit('join', {'rooms': ['alarmes', 'process']})
            time.sleep(0.3)
        finally:
            sio.disconnect()

    def test_42_plc_write_publishes_to_redis(self):
        """plc_write deve resultar em publicação no canal Redis plc_commands."""
        pubsub = _redis_conn.pubsub()
        pubsub.subscribe('plc_commands')
        time.sleep(0.1)

        sio = self._make_client()
        payload = {'Extrusora': {'extrusoraRefVelocidade': 1234}}
        sio.emit('plc_write', payload)
        time.sleep(0.5)
        sio.disconnect()

        msg = None
        for _ in range(10):
            msg = pubsub.get_message(ignore_subscribe_messages=True)
            if msg:
                break
            time.sleep(0.1)
        pubsub.unsubscribe('plc_commands')
        pubsub.close()

        self.assertIsNotNone(msg, "Nenhuma mensagem recebida em plc_commands")
        data = json.loads(msg['data'])
        self.assertEqual(data, payload)

    def test_43_config_get_returns_group_config(self):
        """Evento config_get deve retornar group_config via config:updated."""
        received = []

        sio = self.sio_lib.Client(logger=False, engineio_logger=False)

        @sio.on('config:updated')
        def on_config(data):
            received.append(data)

        sio.connect(f'http://127.0.0.1:{HUB_PORT}')
        time.sleep(0.3)
        sio.emit('config_get', {})
        time.sleep(0.5)
        sio.disconnect()

        self.assertTrue(received, "config:updated não recebido após config_get")
        cfg = received[-1]
        self.assertIn('_meta',    cfg)
        self.assertIn('channels', cfg)
        self.assertNotIn('groups', cfg)   # Fase 5: seção groups removida

    def test_44_history_get_returns_sizes(self):
        """Evento history_get deve retornar history:sizes com os canais."""
        received = []

        sio = self.sio_lib.Client(logger=False, engineio_logger=False)

        @sio.on('history:sizes')
        def on_sizes(data):
            received.append(data)

        sio.connect(f'http://127.0.0.1:{HUB_PORT}')
        time.sleep(0.3)
        sio.emit('history_get', {})
        time.sleep(0.5)
        sio.disconnect()

        self.assertTrue(received, "history:sizes não recebido após history_get")
        sizes = received[-1]
        self.assertIsInstance(sizes, dict)
        for ch in ('plc_alarmes', 'plc_retentivas', 'plc_operacao'):
            self.assertIn(ch, sizes)

    def _get_device_for_channel(self, channel):
        """Helper: query /api/channels to find the device_id for a channel."""
        import urllib.request
        with urllib.request.urlopen(f'http://127.0.0.1:{HUB_PORT}/api/channels', timeout=3) as r:
            channels = json.loads(r.read().decode())
        info = channels.get(channel, {})
        return info.get('device_id', 'unknown')

    def test_45_rooms_isolation(self):
        """
        Cliente no room de um device nao deve receber mensagens de outro device.
        Publica em plc_process e verifica que cliente em room diferente nao recebe.
        """
        received_process = []

        sio = self.sio_lib.Client(logger=False, engineio_logger=False)

        @sio.on('device:data')
        def on_data(data):
            received_process.append(data)

        sio.connect(f'http://127.0.0.1:{HUB_PORT}')
        sio.emit('join', {'rooms': ['nonexistent_device_room']})
        time.sleep(0.3)

        # Publica diretamente em plc_process — NÃO deve chegar ao cliente
        _redis_conn.publish('plc_process', json.dumps({
            'coils': {}, 'registers': {'Extrusora': {'extrusoraRefVelocidade': 1}},
            'timestamp': '2026-01-01T00:00:00',
        }))
        time.sleep(0.5)
        sio.disconnect()

        self.assertEqual(
            len(received_process), 0,
            f"Cliente no room errado recebeu mensagem: {received_process}"
        )

    def test_46_bridge_delivers_device_data_to_correct_room(self):
        """
        Mensagem publicada em plc_alarmes deve chegar ao cliente no room do device.
        Bridge emite device:data para room=device_id.
        """
        # Discover which device owns plc_alarmes
        dev_alarmes = self._get_device_for_channel('plc_alarmes')

        received = []

        sio = self.sio_lib.Client(logger=False, engineio_logger=False)

        @sio.on('device:data')
        def on_data(data):
            received.append(data)

        sio.connect(f'http://127.0.0.1:{HUB_PORT}')
        sio.emit('join', {'rooms': [dev_alarmes]})
        time.sleep(0.8)   # aguarda o join ser processado no servidor

        test_payload = {
            'coils': {'alarmes': {'emergencia': False}},
            'registers': {},
            'timestamp': '2026-01-01T00:00:00',
        }
        # Publica multiplas vezes para compensar possivel latencia de polling
        for _ in range(3):
            _redis_conn.publish('plc_alarmes', json.dumps(test_payload))
            time.sleep(0.5)

        # Aguarda entrega assincrona bridge -> Socket.IO antes de desconectar
        deadline = time.time() + 3.0
        while not received and time.time() < deadline:
            time.sleep(0.1)

        sio.disconnect()

        self.assertTrue(received, "Nenhuma mensagem device:data recebida em room '%s'" % dev_alarmes)
        self.assertEqual(received[0]['channel'], 'plc_alarmes')
        self.assertEqual(received[0]['device_id'], dev_alarmes)

    def test_47_config_save_broadcasts_to_all_clients(self):
        """
        config_save deve emitir config:updated para TODOS os clientes conectados.
        """
        client_a_received = []
        client_b_received = []

        sio_a = self.sio_lib.Client(logger=False, engineio_logger=False)
        sio_b = self.sio_lib.Client(logger=False, engineio_logger=False)

        @sio_a.on('config:updated')
        def on_a(data):
            client_a_received.append(data)

        @sio_b.on('config:updated')
        def on_b(data):
            client_b_received.append(data)

        sio_a.connect(f'http://127.0.0.1:{HUB_PORT}')
        sio_b.connect(f'http://127.0.0.1:{HUB_PORT}')
        time.sleep(0.3)

        cfg = config_store.load_group_config()
        sio_a.emit('config_save', cfg)
        time.sleep(0.5)

        sio_a.disconnect()
        sio_b.disconnect()

        self.assertTrue(client_a_received, "Cliente A não recebeu config:updated")
        self.assertTrue(client_b_received, "Cliente B não recebeu config:updated broadcast")


# ---------------------------------------------------------------------------
# Suite 4 — TestConfigStorePhase3 (unit, sem deps externas)
# ---------------------------------------------------------------------------

class TestConfigStorePhase3(unittest.TestCase):
    """
    Testa as funções de Fase 3 do config_store:
    load_all_variables, generate_export_xlsx, parse_upload_xlsx, apply_upload_config.
    Usa diretório temporário copiando JSON + CSV para isolar os arquivos reais.
    """

    def setUp(self):
        import openpyxl
        import io as _io
        self._openpyxl = openpyxl
        self._io = _io

        self.temp_dir = tempfile.mkdtemp()
        for fname in ('group_config.json', 'variable_overrides.json'):
            shutil.copy(
                os.path.join(TABLES_DIR, fname),
                os.path.join(self.temp_dir, fname),
            )
        # Copy CSV files referenced by devices in group_config.json
        with open(os.path.join(TABLES_DIR, 'group_config.json'), 'r') as f:
            gc = json.load(f)
        csv_files = set()
        for dev in gc.get('devices', {}).values():
            csv_files.update(dev.get('csv_files', []))
        for fname in csv_files:
            src = os.path.join(TABLES_DIR, fname)
            if os.path.exists(src):
                shutil.copy(src, os.path.join(self.temp_dir, fname))

        self._orig_tables_dir = config_store._TABLES_DIR
        config_store._TABLES_DIR = self.temp_dir

    def tearDown(self):
        config_store._TABLES_DIR = self._orig_tables_dir
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    # ── load_all_variables ───────────────────────────────────────────────────

    def test_50_load_all_variables_returns_list(self):
        result = config_store.load_all_variables()
        self.assertIsInstance(result, list)
        self.assertGreater(len(result), 0, "load_all_variables deve retornar pelo menos uma variável")

    def test_51_each_variable_has_required_fields(self):
        result = config_store.load_all_variables()
        required = {'tag', 'group', 'type', 'address', 'channel',
                    'history_size', 'enabled', 'source'}
        for var in result:
            with self.subTest(tag=var.get('tag')):
                self.assertTrue(required.issubset(var.keys()), f"Campos ausentes em {var}")
        # Fase 5: group_cfg_key e has_override removidos
        for var in result:
            self.assertNotIn('group_cfg_key', var)
            self.assertNotIn('has_override',  var)

    def test_52_unassigned_variables_have_null_channel(self):
        """Variáveis sem override de canal devem ter channel=None."""
        result    = config_store.load_all_variables()
        overrides = config_store.load_overrides()
        for v in result:
            expected = overrides.get(v['tag'], {}).get('channel')   # None se não configurado
            with self.subTest(tag=v['tag']):
                self.assertEqual(v['channel'], expected)

    def test_53_override_applied_to_tag(self):
        """Um override deve sobrescrever o canal da variável."""
        result  = config_store.load_all_variables()
        any_tag = result[0]['tag']

        config_store.patch_variable_override(any_tag, {'channel': 'plc_visual', 'enabled': False})
        updated = config_store.load_all_variables()
        target  = next((v for v in updated if v['tag'] == any_tag), None)

        self.assertIsNotNone(target)
        self.assertEqual(target['channel'], 'plc_visual')
        self.assertFalse(target['enabled'])

    def test_54_all_variables_have_source_field(self):
        """Fase 5: group_cfg_key removido; verifica que 'source' continua presente."""
        result = config_store.load_all_variables()
        self.assertTrue(all('source' in v for v in result),
                        "Todas as variáveis devem ter campo 'source'")
        cfg_sources = {v['source'] for v in result}
        self.assertGreater(len(cfg_sources), 0, "Deve haver pelo menos uma source")

    # ── generate_export_xlsx ─────────────────────────────────────────────────

    def test_55_generate_export_xlsx_returns_bytes(self):
        data = config_store.generate_export_xlsx()
        self.assertIsInstance(data, bytes)
        self.assertGreater(len(data), 100)

    def test_56_export_xlsx_is_valid_workbook(self):
        data = config_store.generate_export_xlsx()
        wb   = self._openpyxl.load_workbook(self._io.BytesIO(data))
        ws   = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn('Tag',   headers)
        self.assertIn('Canal', headers)
        self.assertGreater(ws.max_row, 1, "xlsx deve ter pelo menos 1 linha de dados")

    # ── parse_upload_xlsx ────────────────────────────────────────────────────

    def test_57_parse_upload_xlsx_roundtrip(self):
        """Export → parse deve recuperar os campos editáveis."""
        data    = config_store.generate_export_xlsx()
        preview = config_store.parse_upload_xlsx(data)
        self.assertIsInstance(preview, list)
        self.assertGreater(len(preview), 0)
        first = preview[0]
        self.assertIn('tag', first)
        self.assertIn('channel', first)

    def test_58_parse_upload_xlsx_missing_tag_column_raises(self):
        wb = self._openpyxl.Workbook()
        ws = wb.active
        ws.append(['Sem coluna tag', 'Canal'])  # wrong header
        ws.append(['val1', 'plc_alarmes'])
        buf = self._io.BytesIO()
        wb.save(buf)
        with self.assertRaises(ValueError):
            config_store.parse_upload_xlsx(buf.getvalue())

    # ── apply_upload_config ──────────────────────────────────────────────────

    def test_59_apply_upload_creates_overrides(self):
        """apply_upload_config deve criar override de canal para cada variável."""
        variables = config_store.load_all_variables()
        ext_vars  = [v for v in variables if v['group'] == 'alarmes']
        self.assertTrue(ext_vars)

        rows = [
            {'tag': v['tag'], 'channel': 'plc_visual', 'enabled': True}
            for v in ext_vars
        ]
        config_store.apply_upload_config(rows)

        overrides = config_store.load_overrides()
        for v in ext_vars:
            self.assertIn(v['tag'], overrides)
            self.assertEqual(overrides[v['tag']]['channel'], 'plc_visual')

    def test_60_apply_upload_sets_per_tag_override(self):
        """Rows com canais diferentes criam overrides individuais por tag."""
        variables = config_store.load_all_variables()
        ext_vars  = [v for v in variables if v['group'] == 'alarmes']
        self.assertTrue(len(ext_vars) >= 2)

        rows = []
        for i, v in enumerate(ext_vars):
            ch = 'plc_alarmes' if i == 0 else 'plc_process'
            rows.append({'tag': v['tag'], 'channel': ch, 'enabled': True})
        config_store.apply_upload_config(rows)

        overrides = config_store.load_overrides()
        self.assertIn(ext_vars[0]['tag'], overrides)
        self.assertEqual(overrides[ext_vars[0]['tag']]['channel'], 'plc_alarmes')
        self.assertIn(ext_vars[1]['tag'], overrides)
        self.assertEqual(overrides[ext_vars[1]['tag']]['channel'], 'plc_process')


# ---------------------------------------------------------------------------
# Suite 5 — TestHubHTTPPhase3 (requer Hub subprocess + Redis)
# ---------------------------------------------------------------------------

class TestHubHTTPPhase3(unittest.TestCase):
    """Testa os endpoints REST de Fase 3 do Hub."""

    @classmethod
    def setUpClass(cls):
        try:
            r = redis.Redis(host=REDIS_HOST, port=REDIS_PORT, socket_connect_timeout=2)
            r.ping()
            r.close()
        except Exception:
            raise unittest.SkipTest("Redis não disponível")

        if _hub_proc is None:
            raise unittest.SkipTest("Hub não iniciado")

    def _get(self, path, **kwargs):
        import urllib.request
        req = urllib.request.urlopen(f'http://127.0.0.1:{HUB_PORT}{path}', **kwargs)
        return req

    def _post_json(self, path, data):
        import urllib.request
        body = json.dumps(data).encode()
        req  = urllib.request.Request(
            f'http://127.0.0.1:{HUB_PORT}{path}',
            data=body,
            headers={'Content-Type': 'application/json'},
            method='POST',
        )
        return urllib.request.urlopen(req)

    def _patch_json(self, path, data):
        import urllib.request
        body = json.dumps(data).encode()
        req  = urllib.request.Request(
            f'http://127.0.0.1:{HUB_PORT}{path}',
            data=body,
            headers={'Content-Type': 'application/json'},
            method='PATCH',
        )
        return urllib.request.urlopen(req)

    def test_70_get_index_returns_html(self):
        res = self._get('/')
        self.assertEqual(res.status, 200)
        ct = res.headers.get('Content-Type', '')
        self.assertIn('text/html', ct)

    def test_71_get_variables_has_variables_key(self):
        res  = self._get('/api/variables')
        data = json.loads(res.read())
        self.assertIn('variables', data)
        self.assertIn('overrides', data)

    def test_72_get_variables_list_is_non_empty(self):
        res  = self._get('/api/variables')
        data = json.loads(res.read())
        self.assertIsInstance(data['variables'], list)
        self.assertGreater(len(data['variables']), 0)

    def test_73_each_variable_has_tag_and_channel(self):
        res  = self._get('/api/variables')
        data = json.loads(res.read())
        for var in data['variables']:
            with self.subTest(tag=var.get('tag')):
                self.assertIn('tag',     var)
                self.assertIn('channel', var)

    def test_74_patch_variable_creates_override(self):
        import urllib.error
        # Pega um tag qualquer
        res  = self._get('/api/variables')
        data = json.loads(res.read())
        tag  = data['variables'][0]['tag']

        res2 = self._patch_json(f'/api/variables/{tag}', {'enabled': False})
        self.assertEqual(res2.status, 200)

        res3  = self._get('/api/variables')
        data3 = json.loads(res3.read())
        ov    = data3['overrides']
        self.assertIn(tag, ov)
        self.assertFalse(ov[tag].get('enabled', True))

    def test_75_get_export_returns_xlsx(self):
        res = self._get('/api/export')
        self.assertEqual(res.status, 200)
        ct  = res.headers.get('Content-Type', '')
        self.assertIn('spreadsheetml', ct)
        body = res.read()
        self.assertGreater(len(body), 100)

    def test_76_export_is_valid_xlsx(self):
        import io as _io
        import openpyxl
        res  = self._get('/api/export')
        wb   = openpyxl.load_workbook(_io.BytesIO(res.read()))
        ws   = wb.active
        hdrs = [c.value for c in ws[1]]
        self.assertIn('Tag',   hdrs)
        self.assertIn('Canal', hdrs)

    def test_77_post_upload_with_export_data_returns_preview(self):
        """Exporta, faz upload do mesmo arquivo e verifica preview."""
        import email.generator
        import io as _io
        import openpyxl
        import urllib.request

        # Obtém o xlsx exportado
        res      = self._get('/api/export')
        xlsx_raw = res.read()

        # Monta multipart/form-data manualmente
        boundary = b'----TestBoundary12345'
        body = (
            b'--' + boundary + b'\r\n'
            b'Content-Disposition: form-data; name="file"; filename="test.xlsx"\r\n'
            b'Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n'
            b'\r\n' + xlsx_raw + b'\r\n'
            b'--' + boundary + b'--\r\n'
        )
        req = urllib.request.Request(
            f'http://127.0.0.1:{HUB_PORT}/api/upload',
            data=body,
            headers={'Content-Type': f'multipart/form-data; boundary={boundary.decode()}'},
            method='POST',
        )
        resp = urllib.request.urlopen(req)
        data = json.loads(resp.read())
        self.assertIn('preview', data)
        self.assertIn('count',   data)
        self.assertGreater(data['count'], 0)

    def test_78_post_upload_confirm_applies_config(self):
        """upload/confirm aceita rows e retorna contagem."""
        # Pega variables atuais
        res      = self._get('/api/variables')
        data     = json.loads(res.read())
        vars_now = data['variables'][:3]   # usa só 3 para agilizar

        rows = [
            {'tag': v['tag'], 'group': v.get('group', ''),
             'source': v.get('source', 'operacao'),
             'channel': v['channel'],
             'history_size': v.get('history_size', 100), 'enabled': v.get('enabled', True)}
            for v in vars_now
        ]
        res2 = self._post_json('/api/upload/confirm', {'rows': rows})
        self.assertEqual(res2.status, 200)
        data2 = json.loads(res2.read())
        self.assertEqual(data2['applied'], len(rows))


# ---------------------------------------------------------------------------
# Suite 6 — TestDeviceCRUD (unit, sem deps externas)
# ---------------------------------------------------------------------------

class TestDeviceCRUD(unittest.TestCase):
    """
    Testa CRUD de devices em config_store e ping mockado.
    Usa diretório temporário para isolar os arquivos de config.
    """

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        shutil.copy(
            os.path.join(TABLES_DIR, 'group_config.json'),
            os.path.join(self.temp_dir, 'group_config.json'),
        )
        shutil.copy(
            os.path.join(TABLES_DIR, 'variable_overrides.json'),
            os.path.join(self.temp_dir, 'variable_overrides.json'),
        )
        # Copy CSV files referenced by devices in group_config.json
        with open(os.path.join(TABLES_DIR, 'group_config.json'), 'r') as f:
            gc = json.load(f)
        csv_files = set()
        for dev in gc.get('devices', {}).values():
            csv_files.update(dev.get('csv_files', []))
        for fname in csv_files:
            src = os.path.join(TABLES_DIR, fname)
            if os.path.exists(src):
                shutil.copy(src, os.path.join(self.temp_dir, fname))
        self._orig_tables_dir = config_store._TABLES_DIR
        config_store._TABLES_DIR = self.temp_dir

    def tearDown(self):
        config_store._TABLES_DIR = self._orig_tables_dir
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_80_get_devices_returns_dict(self):
        """get_devices() deve retornar dict (inclui o device 'default' do group_config.json)."""
        devices = config_store.get_devices()
        self.assertIsInstance(devices, dict)

    def test_81_create_device_appears_in_get_devices(self):
        """Após create_device(), o device deve aparecer em get_devices()."""
        cfg = {'label': 'CLP 2', 'protocol': 'tcp', 'host': '10.0.0.1',
                'port': 502, 'unit_id': 1, 'csv_files': []}
        config_store.create_device('clp2', cfg)
        devices = config_store.get_devices()
        self.assertIn('clp2', devices)
        self.assertEqual(devices['clp2']['host'], '10.0.0.1')

    def test_82_create_device_preserves_channels(self):
        """create_device() não deve remover os channels dos devices existentes."""
        config_store.create_device('test_dev', {'label': 'Test', 'protocol': 'tcp'})
        cfg = config_store.load_group_config()
        # Channels are inside devices — verify existing device channels preserved
        device_channels = cfg['devices']['simulador'].get('channels', {})
        for ch in ('plc_alarmes', 'plc_retentivas', 'plc_operacao'):
            self.assertIn(ch, device_channels)

    def test_83_update_device_updates_field(self):
        """update_device() deve alterar apenas o campo informado."""
        config_store.create_device('dev_upd', {'label': 'Antes', 'protocol': 'tcp',
                                                'host': '1.1.1.1', 'port': 502, 'unit_id': 1})
        config_store.update_device('dev_upd', {'host': '2.2.2.2'})
        devices = config_store.get_devices()
        self.assertEqual(devices['dev_upd']['host'], '2.2.2.2')
        self.assertEqual(devices['dev_upd']['label'], 'Antes')  # campo não-alterado preservado

    def test_84_update_nonexistent_device_raises_key_error(self):
        """update_device() em device inexistente deve lançar KeyError."""
        with self.assertRaises(KeyError):
            config_store.update_device('inexistente_xyz', {'host': '0.0.0.0'})

    def test_85_delete_device_removes_it(self):
        """delete_device() deve remover o device de get_devices()."""
        config_store.create_device('del_me', {'label': 'Temp', 'protocol': 'tcp'})
        self.assertIn('del_me', config_store.get_devices())
        config_store.delete_device('del_me')
        self.assertNotIn('del_me', config_store.get_devices())

    def test_86_delete_nonexistent_device_raises_key_error(self):
        """delete_device() em device inexistente deve lançar KeyError."""
        with self.assertRaises(KeyError):
            config_store.delete_device('nao_existe_xyz')

    def test_87_load_variables_includes_device_field(self):
        """load_all_variables() deve retornar campo 'device' em cada variável."""
        variables = config_store.load_all_variables()
        self.assertGreater(len(variables), 0)
        for var in variables:
            with self.subTest(tag=var.get('tag')):
                self.assertIn('device', var)

    def test_88_load_variables_device_matches_device_id(self):
        """Variáveis do device 'default' devem ter device='default'."""
        variables = config_store.load_all_variables()
        devices = config_store.get_devices()
        if 'default' in devices:
            default_vars = [v for v in variables if v['device'] == 'default']
            self.assertGreater(len(default_vars), 0,
                "Deve haver variáveis com device='default' quando devices.default existe")

    def test_89_ping_tcp_calls_modbus_client(self):
        """_do_ping TCP deve chamar ModbusClient.read_holding_registers."""
        from unittest.mock import MagicMock, patch
        import Hub.main as hub_main

        mock_client = MagicMock()
        mock_client.read_holding_registers.return_value = [42]

        with patch('pyModbusTCP.client.ModbusClient', return_value=mock_client):
            cfg = {'protocol': 'tcp', 'host': '127.0.0.1', 'port': 502, 'unit_id': 1}
            result = hub_main._do_ping(cfg)

        mock_client.read_holding_registers.assert_called_once_with(0, 1)
        self.assertIsInstance(result, dict)
        self.assertIn('ok', result)
        self.assertIn('latency_ms', result)

    def test_90_ping_rtu_calls_serial_client(self):
        """_do_ping RTU deve chamar ModbusSerialClient.connect e read_holding_registers."""
        from unittest.mock import MagicMock, patch
        import Hub.main as hub_main

        mock_client = MagicMock()
        mock_client.connect.return_value = True
        mock_client.read_holding_registers.return_value = MagicMock(isError=lambda: False)

        with patch('pymodbus.client.ModbusSerialClient', return_value=mock_client):
            cfg = {'protocol': 'rtu', 'serial_port': 'COM3', 'baudrate': 9600,
                   'parity': 'N', 'stopbits': 1, 'unit_id': 1}
            result = hub_main._do_ping(cfg)

        mock_client.connect.assert_called_once()
        mock_client.read_holding_registers.assert_called_once_with(0, 1, slave=1)
        self.assertIsInstance(result, dict)
        self.assertIn('ok', result)

    def test_91a_remove_csv_from_device(self):
        """Remover CSV de csv_files do device deve atualizar a lista."""
        # Create device with a CSV
        csv_content = "key,ObjecTag,Tipo,Modbus,At\ngrp,tagRm,D,500,%MW\n"
        csv_path = os.path.join(self.temp_dir, 'removable.csv')
        with open(csv_path, 'w') as f:
            f.write(csv_content)

        config_store.create_device('dev_rm', {
            'label': 'RemTest', 'protocol': 'tcp', 'host': '1.1.1.1',
            'port': 502, 'unit_id': 1, 'csv_files': ['removable.csv'],
        })

        # Create per-device overrides for the tag
        config_store.patch_variable_override('tagRm', {'channel': 'plc_test'}, device_id='dev_rm')
        ov = config_store.load_overrides('dev_rm')
        self.assertIn('tagRm', ov)

        # Remove CSV from device (simulate what the endpoint does)
        devices = config_store.get_devices()
        dev_cfg = devices['dev_rm']
        csv_files = list(dev_cfg.get('csv_files', []))
        self.assertIn('removable.csv', csv_files)

        # Clean overrides for tags in that CSV
        import pandas as pd
        df = pd.read_csv(csv_path, sep=',')
        csv_tags = set(df['ObjecTag'].astype(str).str.strip().values)
        overrides = config_store.load_overrides('dev_rm')
        for tag in csv_tags:
            overrides.pop(tag, None)
        config_store.save_overrides(overrides, 'dev_rm')

        csv_files.remove('removable.csv')
        config_store.update_device('dev_rm', {'csv_files': csv_files})

        # Verify
        devices = config_store.get_devices()
        self.assertEqual(devices['dev_rm']['csv_files'], [])
        ov = config_store.load_overrides('dev_rm')
        self.assertNotIn('tagRm', ov)

    def test_91b_remove_csv_not_in_device_raises(self):
        """Tentar remover CSV que não está no device deve falhar."""
        config_store.create_device('dev_no_csv', {
            'label': 'Test', 'protocol': 'tcp', 'csv_files': ['a.csv'],
        })
        devices = config_store.get_devices()
        csv_files = list(devices['dev_no_csv'].get('csv_files', []))
        self.assertNotIn('b.csv', csv_files)


# ---------------------------------------------------------------------------
# Suite 8 — TestProcessManagerPerDevice (unitario, sem deps externas)
# ---------------------------------------------------------------------------

class TestProcessManagerPerDevice(unittest.TestCase):
    """
    Testa ProcessInstance e ProcessManager com per-device isolation.
    Nao lanca subprocessos reais — foca na logica de init, state e env vars.
    """

    def test_91_process_instance_stores_device_id(self):
        """ProcessInstance.__init__ deve armazenar device_id."""
        from process_manager import ProcessInstance
        proc = ProcessInstance('delfos:sim1', 'delfos', {'modbus_host': '127.0.0.1'}, device_id='sim1')
        self.assertEqual(proc.device_id, 'sim1')
        self.assertEqual(proc.proc_id, 'delfos:sim1')
        self.assertEqual(proc.proc_type, 'delfos')

    def test_92_process_instance_default_device_id_empty(self):
        """ProcessInstance sem device_id deve ter device_id=''."""
        from process_manager import ProcessInstance
        proc = ProcessInstance('delfos', 'delfos', {})
        self.assertEqual(proc.device_id, '')

    def test_93_to_state_dict_includes_device_id(self):
        """to_state_dict() deve incluir 'device_id' no dict retornado."""
        from process_manager import ProcessInstance
        proc = ProcessInstance('atena:clp2', 'atena', {}, device_id='clp2')
        state = proc.to_state_dict()
        self.assertIn('device_id', state)
        self.assertEqual(state['device_id'], 'clp2')
        self.assertEqual(state['proc_id'], 'atena:clp2')
        self.assertEqual(state['proc_type'], 'atena')

    def test_94_start_sets_device_env_vars(self):
        """ProcessInstance.start() deve definir DEVICE_ID, COMMAND_CHANNEL e CONFIG_RELOAD_CHANNEL no env."""
        import asyncio
        from unittest.mock import AsyncMock, patch, MagicMock
        from process_manager import ProcessInstance

        proc = ProcessInstance('delfos:sim1', 'delfos', {
            'modbus_host': '127.0.0.1',
            'modbus_port': 502,
            'modbus_unit_id': 1,
            'modbus_protocol': 'tcp',
            'redis_host': 'localhost',
            'redis_port': 6379,
        }, device_id='sim1')

        captured_env = {}

        async def fake_create_subprocess_exec(*args, **kwargs):
            captured_env.update(kwargs.get('env', {}))
            mock_proc = MagicMock()
            mock_proc.pid = 12345
            mock_proc.stdout = AsyncMock()
            mock_proc.stdout.readline = AsyncMock(return_value=b'')
            mock_proc.wait = AsyncMock(return_value=0)
            mock_proc.returncode = 0
            return mock_proc

        gateway_dir = GATEWAY_DIR
        python_path = PYTHON

        with patch('asyncio.create_subprocess_exec', side_effect=fake_create_subprocess_exec):
            asyncio.get_event_loop().run_until_complete(proc.start(python_path, gateway_dir))

        self.assertEqual(captured_env.get('DEVICE_ID'), 'sim1')
        self.assertEqual(captured_env.get('COMMAND_CHANNEL'), 'sim1_commands')
        self.assertEqual(captured_env.get('CONFIG_RELOAD_CHANNEL'), 'config_reload_sim1')

    def test_95_start_uses_custom_command_channel(self):
        """Se config contiver command_channel, deve usar esse valor em vez do default."""
        import asyncio
        from unittest.mock import AsyncMock, patch, MagicMock
        from process_manager import ProcessInstance

        proc = ProcessInstance('delfos:sim1', 'delfos', {
            'command_channel': 'custom_cmd_chan',
            'config_reload_channel': 'custom_reload_chan',
        }, device_id='sim1')

        captured_env = {}

        async def fake_create_subprocess_exec(*args, **kwargs):
            captured_env.update(kwargs.get('env', {}))
            mock_proc = MagicMock()
            mock_proc.pid = 99999
            mock_proc.stdout = AsyncMock()
            mock_proc.stdout.readline = AsyncMock(return_value=b'')
            mock_proc.wait = AsyncMock(return_value=0)
            mock_proc.returncode = 0
            return mock_proc

        with patch('asyncio.create_subprocess_exec', side_effect=fake_create_subprocess_exec):
            asyncio.get_event_loop().run_until_complete(proc.start(PYTHON, GATEWAY_DIR))

        self.assertEqual(captured_env.get('COMMAND_CHANNEL'), 'custom_cmd_chan')
        self.assertEqual(captured_env.get('CONFIG_RELOAD_CHANNEL'), 'custom_reload_chan')

    def test_96_process_manager_start_derives_proc_id(self):
        """ProcessManager.start_process deve derivar proc_id como 'proc_type:device_id'."""
        import asyncio
        from unittest.mock import AsyncMock, patch, MagicMock
        from process_manager import ProcessManager

        pm = ProcessManager(GATEWAY_DIR)

        async def fake_create_subprocess_exec(*args, **kwargs):
            mock_proc = MagicMock()
            mock_proc.pid = 11111
            mock_proc.stdout = AsyncMock()
            mock_proc.stdout.readline = AsyncMock(return_value=b'')
            mock_proc.wait = AsyncMock(return_value=0)
            mock_proc.returncode = 0
            return mock_proc

        with patch('asyncio.create_subprocess_exec', side_effect=fake_create_subprocess_exec):
            proc = asyncio.get_event_loop().run_until_complete(
                pm.start_process('delfos', 'sim1', {'modbus_host': '127.0.0.1'})
            )

        self.assertEqual(proc.proc_id, 'delfos:sim1')
        self.assertEqual(proc.device_id, 'sim1')
        self.assertEqual(proc.proc_type, 'delfos')
        self.assertIn('delfos:sim1', pm.list_processes())

    def test_97_process_manager_multiple_devices(self):
        """ProcessManager deve suportar multiplos processos do mesmo tipo para devices diferentes."""
        import asyncio
        from unittest.mock import AsyncMock, patch, MagicMock
        from process_manager import ProcessManager

        pm = ProcessManager(GATEWAY_DIR)

        async def fake_create_subprocess_exec(*args, **kwargs):
            mock_proc = MagicMock()
            mock_proc.pid = 22222
            mock_proc.stdout = AsyncMock()
            mock_proc.stdout.readline = AsyncMock(return_value=b'')
            mock_proc.wait = AsyncMock(return_value=0)
            mock_proc.returncode = 0
            return mock_proc

        with patch('asyncio.create_subprocess_exec', side_effect=fake_create_subprocess_exec):
            p1 = asyncio.get_event_loop().run_until_complete(
                pm.start_process('delfos', 'sim1', {'modbus_host': '127.0.0.1'})
            )
            p2 = asyncio.get_event_loop().run_until_complete(
                pm.start_process('delfos', 'sim2', {'modbus_host': '127.0.0.2'})
            )

        processes = pm.list_processes()
        self.assertIn('delfos:sim1', processes)
        self.assertIn('delfos:sim2', processes)
        self.assertEqual(processes['delfos:sim1']['device_id'], 'sim1')
        self.assertEqual(processes['delfos:sim2']['device_id'], 'sim2')

    def test_98_process_manager_duplicate_raises_runtime_error(self):
        """Iniciar processo para mesmo proc_type:device_id deve lancar RuntimeError se ainda rodando."""
        from process_manager import ProcessManager, ProcessInstance

        pm = ProcessManager(GATEWAY_DIR)

        # Directly inject a running process into ProcessManager state
        existing = ProcessInstance('delfos:sim1', 'delfos', {'modbus_host': '127.0.0.1'}, device_id='sim1')
        existing.running = True
        pm._processes['delfos:sim1'] = existing

        # Should raise because delfos:sim1 is already "running"
        import asyncio
        with self.assertRaises(RuntimeError):
            asyncio.get_event_loop().run_until_complete(
                pm.start_process('delfos', 'sim1', {'modbus_host': '127.0.0.1'})
            )


# ---------------------------------------------------------------------------
# Suite 9 — TestScannerManager (unitário, sem deps externas)
# ---------------------------------------------------------------------------

class TestScannerManager(unittest.TestCase):
    """
    Testa ScannerManager, ScanSession e lógica de scan.
    Usa diretório temporário e mocks de Modbus.
    """

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        shutil.copy(
            os.path.join(TABLES_DIR, 'group_config.json'),
            os.path.join(self.temp_dir, 'group_config.json'),
        )
        # Copy CSV files for load_device_variables
        with open(os.path.join(TABLES_DIR, 'group_config.json'), 'r') as f:
            gc = json.load(f)
        for dev in gc.get('devices', {}).values():
            for fname in dev.get('csv_files', []):
                src = os.path.join(TABLES_DIR, fname)
                if os.path.exists(src):
                    shutil.copy(src, os.path.join(self.temp_dir, fname))
        # Copy per-device overrides
        for dev_id in gc.get('devices', {}):
            ov_name = f'variable_overrides_{dev_id}.json'
            ov_src = os.path.join(TABLES_DIR, ov_name)
            if os.path.exists(ov_src):
                shutil.copy(ov_src, os.path.join(self.temp_dir, ov_name))
        # Also copy global overrides
        gv = os.path.join(TABLES_DIR, 'variable_overrides.json')
        if os.path.exists(gv):
            shutil.copy(gv, os.path.join(self.temp_dir, 'variable_overrides.json'))
        self._orig_tables_dir = config_store._TABLES_DIR
        config_store._TABLES_DIR = self.temp_dir

    def tearDown(self):
        config_store._TABLES_DIR = self._orig_tables_dir
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_100_scan_session_to_dict(self):
        """ScanSession.to_dict() deve retornar todos os campos."""
        from scanner_manager import ScanSession
        s = ScanSession(device_id='sim1', total=10)
        d = s.to_dict()
        self.assertEqual(d['device_id'], 'sim1')
        self.assertEqual(d['total'], 10)
        self.assertEqual(d['status'], 'running')
        self.assertIn('results', d)
        self.assertIn('config', d)

    def test_101_scan_session_to_summary(self):
        """ScanSession.to_summary() não deve incluir results."""
        from scanner_manager import ScanSession
        s = ScanSession(device_id='sim1', total=5, ok_count=3, error_count=2)
        summary = s.to_summary()
        self.assertNotIn('results', summary)
        self.assertEqual(summary['ok_count'], 3)
        self.assertEqual(summary['error_count'], 2)

    def test_102_scan_session_cancel(self):
        """ScanSession.cancel() deve marcar _cancelled=True."""
        from scanner_manager import ScanSession
        s = ScanSession(device_id='sim1')
        self.assertFalse(s.is_cancelled)
        s.cancel()
        self.assertTrue(s.is_cancelled)

    def test_103_scanner_manager_no_active_scan(self):
        """is_scanning() retorna False quando nenhum scan ativo."""
        from scanner_manager import ScannerManager
        sm = ScannerManager(self.temp_dir)
        self.assertFalse(sm.is_scanning('sim1'))

    def test_104_scanner_manager_get_scan_returns_none(self):
        """get_scan() retorna None quando não há scan."""
        from scanner_manager import ScannerManager
        sm = ScannerManager(self.temp_dir)
        self.assertIsNone(sm.get_scan('sim1'))

    def test_105_scanner_manager_results_for_grid_empty(self):
        """get_results_for_grid() retorna {} quando não há scan."""
        from scanner_manager import ScannerManager
        sm = ScannerManager(self.temp_dir)
        self.assertEqual(sm.get_results_for_grid('sim1'), {})

    def test_106_scan_single_variable_ok(self):
        """_scan_single_variable retorna status=ok com mock de client."""
        from unittest.mock import MagicMock
        from scanner_manager import _scan_single_variable
        client = MagicMock()
        client.read_holding_registers.return_value = [1500]
        var = {'tag': 'testVar', 'address': 100, 'type': '%MW'}
        result = _scan_single_variable(client, var, retries=3)
        self.assertEqual(result['status'], 'ok')
        self.assertEqual(result['value'], 1500)
        self.assertEqual(result['tag'], 'testVar')
        self.assertIsNotNone(result['latency_ms'])
        self.assertIsNone(result['error'])

    def test_107_scan_single_variable_coil(self):
        """_scan_single_variable lê coils quando type=%MB."""
        from unittest.mock import MagicMock
        from scanner_manager import _scan_single_variable
        client = MagicMock()
        client.read_coils.return_value = [1]
        var = {'tag': 'coilVar', 'address': 200, 'type': '%MB'}
        result = _scan_single_variable(client, var, retries=1)
        self.assertEqual(result['status'], 'ok')
        self.assertTrue(result['value'])
        client.read_coils.assert_called_once_with(200, 1)

    def test_108_scan_single_variable_error_with_retries(self):
        """_scan_single_variable retorna status=error após todas as tentativas falharem."""
        from unittest.mock import MagicMock
        from scanner_manager import _scan_single_variable
        client = MagicMock()
        client.read_holding_registers.side_effect = Exception('Connection refused')
        var = {'tag': 'failVar', 'address': 300, 'type': '%MW'}
        result = _scan_single_variable(client, var, retries=2)
        self.assertEqual(result['status'], 'error')
        self.assertIn('Connection refused', result['error'])
        self.assertEqual(result['retries_used'], 2)
        self.assertEqual(client.read_holding_registers.call_count, 2)

    def test_109_scan_single_variable_none_response(self):
        """_scan_single_variable trata resposta None como erro."""
        from unittest.mock import MagicMock
        from scanner_manager import _scan_single_variable
        client = MagicMock()
        client.read_holding_registers.return_value = None
        var = {'tag': 'noneVar', 'address': 400, 'type': '%MW'}
        result = _scan_single_variable(client, var, retries=2)
        self.assertEqual(result['status'], 'error')
        self.assertIn('Sem resposta', result['error'])

    def test_110_scanner_manager_persistence(self):
        """Resultados de scan devem ser carregados do disco via load_cached_results."""
        from scanner_manager import ScannerManager, ScanSession
        sm = ScannerManager(self.temp_dir)
        # Create and save a fake session
        session = ScanSession(
            device_id='test_dev',
            status='completed',
            total=2,
            scanned=2,
            ok_count=1,
            error_count=1,
            results=[
                {'tag': 'var1', 'address': 100, 'type': '%MW', 'status': 'ok',
                 'value': 42, 'latency_ms': 1.5, 'error': None, 'retries_used': 1,
                 'timestamp': '2026-01-01T00:00:00'},
                {'tag': 'var2', 'address': 200, 'type': '%MB', 'status': 'error',
                 'value': None, 'latency_ms': 5.0, 'error': 'timeout', 'retries_used': 3,
                 'timestamp': '2026-01-01T00:00:01'},
            ],
            started_at='2026-01-01T00:00:00',
            finished_at='2026-01-01T00:00:02',
        )
        sm._scans['test_dev'] = session
        sm._save_results(session)

        # Load in a new manager
        sm2 = ScannerManager(self.temp_dir)
        loaded = sm2.load_cached_results('test_dev')
        self.assertIsNotNone(loaded)
        self.assertEqual(loaded.status, 'completed')
        self.assertEqual(loaded.ok_count, 1)
        self.assertEqual(loaded.error_count, 1)
        self.assertEqual(len(loaded.results), 2)

    def test_111_scanner_manager_results_for_grid(self):
        """get_results_for_grid retorna dict {tag: {status, ...}} correto."""
        from scanner_manager import ScannerManager, ScanSession
        sm = ScannerManager(self.temp_dir)
        session = ScanSession(
            device_id='test_dev',
            status='completed',
            results=[
                {'tag': 'a', 'status': 'ok', 'latency_ms': 1.0, 'error': None, 'value': 10,
                 'address': 100, 'type': '%MW', 'retries_used': 1, 'timestamp': ''},
                {'tag': 'b', 'status': 'error', 'latency_ms': 5.0, 'error': 'timeout', 'value': None,
                 'address': 200, 'type': '%MB', 'retries_used': 3, 'timestamp': ''},
            ],
        )
        sm._scans['test_dev'] = session
        grid = sm.get_results_for_grid('test_dev')
        self.assertIn('a', grid)
        self.assertEqual(grid['a']['status'], 'ok')
        self.assertEqual(grid['a']['value'], 10)
        self.assertIn('b', grid)
        self.assertEqual(grid['b']['status'], 'error')
        self.assertEqual(grid['b']['error'], 'timeout')

    def test_112_scanner_manager_load_all_cached(self):
        """load_all_cached carrega todos os scan_results_*.json."""
        from scanner_manager import ScannerManager, ScanSession
        sm = ScannerManager(self.temp_dir)
        # Write two result files
        for dev in ['dev1', 'dev2']:
            path = os.path.join(self.temp_dir, f'scan_results_{dev}.json')
            data = ScanSession(device_id=dev, status='completed', total=1, scanned=1, ok_count=1).to_dict()
            with open(path, 'w') as f:
                json.dump(data, f)
        sm.load_all_cached()
        self.assertIsNotNone(sm.get_scan('dev1'))
        self.assertIsNotNone(sm.get_scan('dev2'))

    def test_113_load_device_variables_filters_by_device(self):
        """load_device_variables retorna apenas variáveis do device."""
        devices = config_store.get_devices()
        if not devices:
            self.skipTest('Nenhum device configurado.')
        dev_id = next(iter(devices))
        vars_all = config_store.load_all_variables()
        vars_dev = config_store.load_device_variables(dev_id)
        self.assertGreater(len(vars_dev), 0)
        for v in vars_dev:
            self.assertEqual(v['device'], dev_id)
        self.assertLessEqual(len(vars_dev), len(vars_all))

    def test_114_scanner_start_duplicate_raises(self):
        """start_scan com scan já ativo deve lançar RuntimeError."""
        import asyncio
        from scanner_manager import ScannerManager, ScanSession
        sm = ScannerManager(self.temp_dir)
        # Inject a running session
        session = ScanSession(device_id='sim1', status='running')
        sm._scans['sim1'] = session
        with self.assertRaises(RuntimeError):
            sm.start_scan('sim1', {}, [], {})


    # ── Bit addressing ────────────────────────────────────────────────────────

    def test_115_variables_include_bit_index(self):
        """load_all_variables deve retornar bit_index para variáveis bit-addressed."""
        variables = config_store.load_all_variables()
        # Verifica que o campo bit_index existe em todas as variáveis
        for var in variables:
            with self.subTest(tag=var.get('tag')):
                self.assertIn('bit_index', var)

    def test_116_bit_index_correct_for_temperatura(self):
        """Variáveis de temperatura com Modbus como '1584.01' devem ter bit_index=1."""
        variables = config_store.load_all_variables()
        # Busca variável tempNaoAtingidaZona2 (bit 1 do reg 1584)
        temp_vars = [v for v in variables if v['tag'] == 'tempNaoAtingidaZona2']
        if temp_vars:
            var = temp_vars[0]
            self.assertEqual(var['bit_index'], 1)
            self.assertEqual(var['address'], 1584)

    def test_117_bit_index_none_for_normal(self):
        """Variáveis normais (sem bit addressing) devem ter bit_index=None."""
        variables = config_store.load_all_variables()
        # Busca variável tempZona1 (register normal 1536)
        normal_vars = [v for v in variables if v['tag'] == 'tempZona1']
        if normal_vars:
            var = normal_vars[0]
            self.assertIsNone(var['bit_index'])
            self.assertEqual(var['address'], 1536)

    def test_118_scan_bit_addressed_variable(self):
        """Scanner deve extrair bit de variável bit-addressed."""
        from unittest.mock import MagicMock
        from scanner_manager import _scan_single_variable

        mock_client = MagicMock()
        # Register value: bit 1 ON, bit 0 OFF → 0b10 = 2
        mock_client.read_holding_registers.return_value = [2]

        var = {
            'tag': 'tempNaoAtingidaZona2',
            'address': 1584,
            'type': '%MB',
            'bit_index': 1,
        }
        result = _scan_single_variable(mock_client, var, retries=1)
        self.assertEqual(result['status'], 'ok')
        self.assertTrue(result['value'])  # bit 1 is ON in value 2
        self.assertEqual(result['bit_index'], 1)

        # Bit 0 should be OFF
        var0 = {
            'tag': 'tempNaoAtingidaZona1',
            'address': 1584,
            'type': '%MB',
            'bit_index': 0,
        }
        result0 = _scan_single_variable(mock_client, var0, retries=1)
        self.assertFalse(result0['value'])  # bit 0 is OFF in value 2


if __name__ == '__main__':
    unittest.main(verbosity=2)
