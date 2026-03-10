"""
config_store — leitura e escrita das configurações do gateway.

Gerencia dois arquivos em tables/:
  - group_config.json      : mapeamento grupo → canal + channels section (delay_ms, history_size)
  - variable_overrides.json: exceções por tag (enabled, channel)

O caminho base (_TABLES_DIR) pode ser sobrescrito via variável de ambiente
TABLES_DIR ou diretamente no atributo de módulo (útil em testes).
"""

import io
import json
import logging
import os
import sys

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from shared.bit_addressing import parse_modbus_address

logger = logging.getLogger(__name__)

_HUB_DIR     = os.path.dirname(os.path.abspath(__file__))
_GATEWAY_DIR = os.path.dirname(_HUB_DIR)

def _resolve_tables_dir() -> str:
    """
    Resolve o diretório de tabelas.
    Caminhos relativos (ex: '../tables') são resolvidos a partir de _HUB_DIR,
    não do CWD — garante comportamento correto independente de onde o processo
    é iniciado ou de qual .env foi carregado no ambiente herdado.
    """
    tables_env = os.environ.get('TABLES_DIR')
    if tables_env:
        if not os.path.isabs(tables_env):
            return os.path.normpath(os.path.join(_HUB_DIR, tables_env))
        return tables_env
    return os.path.join(_GATEWAY_DIR, 'tables')

_TABLES_DIR = _resolve_tables_dir()


def _group_config_path() -> str:
    return os.path.join(_TABLES_DIR, 'group_config.json')


def _overrides_path(device_id: str | None = None) -> str:
    if device_id:
        return os.path.join(_TABLES_DIR, f'variable_overrides_{device_id}.json')
    return os.path.join(_TABLES_DIR, 'variable_overrides.json')


# ── Leitura ──────────────────────────────────────────────────────────────────

def load_group_config() -> dict:
    """Carrega group_config.json. Lança exceção se não encontrar."""
    with open(_group_config_path(), 'r', encoding='utf-8') as f:
        return json.load(f)


def load_overrides(device_id: str | None = None) -> dict:
    """Carrega variable_overrides.json (ou per-device). Strip de delay_ms em cada entrada. Retorna {} se ausente."""
    path = _overrides_path(device_id)
    if not os.path.exists(path):
        return {}
    with open(path, 'r', encoding='utf-8') as f:
        raw = json.load(f)
    # Strip delay_ms de todas as entradas (campo removido da camada de variáveis)
    return {tag: {k: v for k, v in cfg.items() if k != 'delay_ms'} for tag, cfg in raw.items()}


# ── Escrita ───────────────────────────────────────────────────────────────────

def save_group_config(config: dict) -> None:
    """Persiste group_config.json."""
    with open(_group_config_path(), 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2, ensure_ascii=False)
    logger.info("group_config.json salvo.")


def save_overrides(overrides: dict, device_id: str | None = None) -> None:
    """Persiste variable_overrides.json (ou per-device). Strip de delay_ms antes de gravar."""
    cleaned = {tag: {k: v for k, v in cfg.items() if k != 'delay_ms'} for tag, cfg in overrides.items()}
    path = _overrides_path(device_id)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(cleaned, f, indent=2, ensure_ascii=False)
    suffix = f" (device={device_id})" if device_id else ""
    logger.info("variable_overrides%s.json salvo.", suffix)


# ── Operações de device ───────────────────────────────────────────────────────

def get_devices() -> dict:
    """Retorna {device_id: cfg} para todos os devices configurados."""
    return load_group_config().get('devices', {})


def create_device(device_id: str, cfg: dict) -> None:
    """Cria ou substitui um device em group_config['devices'].
    Auto-generates 'channels' (empty dict) and 'command_channel' if not provided."""
    config = load_group_config()
    cfg.setdefault('channels', {})
    cfg.setdefault('command_channel', f'{device_id}_commands')
    config.setdefault('devices', {})[device_id] = cfg
    save_group_config(config)
    logger.info("Device '%s' criado.", device_id)


def update_device(device_id: str, fields: dict) -> None:
    """Atualiza campos de um device existente. Lança KeyError se não encontrado."""
    config = load_group_config()
    devices = config.setdefault('devices', {})
    if device_id not in devices:
        raise KeyError(f"Device '{device_id}' não encontrado.")
    devices[device_id].update({k: v for k, v in fields.items() if v is not None})
    save_group_config(config)
    logger.info("Device '%s' atualizado: %s", device_id, fields)


def delete_device(device_id: str) -> None:
    """Remove um device. Lança KeyError se não encontrado."""
    config = load_group_config()
    if device_id not in config.get('devices', {}):
        raise KeyError(f"Device '{device_id}' não encontrado.")
    del config['devices'][device_id]
    save_group_config(config)
    logger.info("Device '%s' removido.", device_id)


# ── Operações de canal ────────────────────────────────────────────────────────

def get_channels() -> dict:
    """
    Retorna {channel: {delay_ms, history_size, device_id}} para todos os canais configurados.
    Fonte: devices[*].channels (per-device), com fallback para seção global 'channels'.
    """
    config = load_group_config()
    meta   = config.get('_meta', {})
    default_delay = meta.get('default_delay_ms', 1000)
    default_hist  = meta.get('default_history_size', 100)

    result: dict = {}

    # Per-device channels (primary source)
    for dev_id, dev_cfg in config.get('devices', {}).items():
        for ch, ch_cfg in dev_cfg.get('channels', {}).items():
            result[ch] = {
                'delay_ms':     ch_cfg.get('delay_ms',    default_delay),
                'history_size': ch_cfg.get('history_size', default_hist),
                'enabled':      ch_cfg.get('enabled', True),
                'device_id':    dev_id,
            }

    # Legacy fallback: global 'channels' section (for backward compat)
    for ch, ch_cfg in config.get('channels', {}).items():
        if ch not in result:
            result[ch] = {
                'delay_ms':     ch_cfg.get('delay_ms',    default_delay),
                'history_size': ch_cfg.get('history_size', default_hist),
            }

    return result


def get_device_channels(device_id: str) -> dict:
    """
    Retorna {channel: {delay_ms, history_size}} apenas para os canais de um device específico.
    """
    config = load_group_config()
    meta   = config.get('_meta', {})
    default_delay = meta.get('default_delay_ms', 1000)
    default_hist  = meta.get('default_history_size', 100)

    dev = config.get('devices', {}).get(device_id, {})
    result: dict = {}
    for ch, ch_cfg in dev.get('channels', {}).items():
        result[ch] = {
            'delay_ms':     ch_cfg.get('delay_ms',    default_delay),
            'history_size': ch_cfg.get('history_size', default_hist),
            'enabled':      ch_cfg.get('enabled', True),
        }
    return result


def get_channel_history_sizes() -> dict:
    """Wrapper backward-compat: retorna {channel: history_size}."""
    return {ch: v['history_size'] for ch, v in get_channels().items()}


def create_channel(channel: str, delay_ms: int = 1000, history_size: int = 100, device_id: str | None = None) -> None:
    """Cria ou atualiza um canal. Se device_id for informado, cria dentro de devices[device_id].channels."""
    config = load_group_config()
    if not device_id:
        raise ValueError("device_id obrigatorio para criar canal.")
    dev = config.get('devices', {}).get(device_id)
    if not dev:
        raise KeyError(f"Device '{device_id}' nao encontrado.")
    dev.setdefault('channels', {})[channel] = {
        'delay_ms': delay_ms,
        'history_size': history_size,
    }
    save_group_config(config)
    logger.info("Canal '%s' criado (delay_ms=%d, history_size=%d, device=%s).", channel, delay_ms, history_size, device_id)


SYSTEM_CHANNELS = frozenset([
    'user_status', 'ia_status', 'ia_data',
])


def delete_channel(channel: str, device_id: str | None = None) -> None:
    """Remove um canal. Se device_id informado, remove de devices[device_id].channels."""
    if channel in SYSTEM_CHANNELS:
        raise ValueError(f"Canal '{channel}' é um canal de sistema e não pode ser removido.")
    config = load_group_config()
    if device_id:
        dev = config.get('devices', {}).get(device_id)
        if not dev:
            raise KeyError(f"Device '{device_id}' nao encontrado.")
        channels = dev.get('channels', {})
        if channel not in channels:
            raise KeyError(f"Canal '{channel}' não encontrado em device '{device_id}'.")
        del channels[channel]
    else:
        # Legacy fallback: global channels section
        channels = config.get('channels', {})
        if channel not in channels:
            raise KeyError(f"Canal '{channel}' não encontrado em channels.")
        del channels[channel]
    save_group_config(config)
    logger.info("Canal '%s' removido (device=%s).", channel, device_id)


def update_channel_delay(channel: str, delay_ms: int, device_id: str | None = None) -> None:
    """Atualiza delay_ms do canal. Se device_id informado, opera em devices[device_id].channels."""
    config = load_group_config()
    if device_id:
        dev = config.get('devices', {}).get(device_id)
        if not dev:
            raise KeyError(f"Device '{device_id}' nao encontrado.")
        dev.setdefault('channels', {}).setdefault(channel, {})['delay_ms'] = delay_ms
    else:
        config.setdefault('channels', {}).setdefault(channel, {})['delay_ms'] = delay_ms
    save_group_config(config)
    logger.info("delay_ms=%d aplicado ao canal '%s' (device=%s).", delay_ms, channel, device_id)


def update_channel_history_size(channel: str, size: int, device_id: str | None = None) -> None:
    """Atualiza history_size do canal. Se device_id informado, opera em devices[device_id].channels."""
    config = load_group_config()
    if device_id:
        dev = config.get('devices', {}).get(device_id)
        if not dev:
            raise KeyError(f"Device '{device_id}' nao encontrado.")
        dev.setdefault('channels', {}).setdefault(channel, {})['history_size'] = size
    else:
        config.setdefault('channels', {}).setdefault(channel, {})['history_size'] = size
    save_group_config(config)
    logger.info("history_size=%d aplicado ao canal '%s' (device=%s).", size, channel, device_id)


def update_channel_enabled(channel: str, enabled: bool, device_id: str | None = None) -> None:
    """Habilita ou desabilita um canal. Se device_id informado, opera em devices[device_id].channels."""
    config = load_group_config()
    if device_id:
        dev = config.get('devices', {}).get(device_id)
        if not dev:
            raise KeyError(f"Device '{device_id}' nao encontrado.")
        ch_entry = dev.get('channels', {}).get(channel)
        if ch_entry is None:
            raise KeyError(f"Canal '{channel}' nao encontrado no device '{device_id}'.")
        ch_entry['enabled'] = enabled
    else:
        ch_entry = config.get('channels', {}).get(channel)
        if ch_entry is None:
            raise KeyError(f"Canal '{channel}' nao encontrado.")
        ch_entry['enabled'] = enabled
    save_group_config(config)
    logger.info("enabled=%s aplicado ao canal '%s' (device=%s).", enabled, channel, device_id)


def find_tag_device(tag: str) -> str | None:
    """Retorna o device_id que contém a tag (via CSVs), ou None se não encontrada."""
    cfg = load_group_config()
    for dev_id, dev_cfg in cfg.get('devices', {}).items():
        for csv_name in dev_cfg.get('csv_files', []):
            csv_path = os.path.join(_TABLES_DIR, csv_name)
            if not os.path.exists(csv_path):
                continue
            try:
                df = pd.read_csv(csv_path, sep=',')
                if 'ObjecTag' in df.columns and tag in df['ObjecTag'].astype(str).str.strip().values:
                    return dev_id
            except Exception:
                continue
    return None


def validate_channel_device(tag: str, channel: str | None, device_id: str | None = None) -> None:
    """Valida que o canal pertence ao mesmo device da tag. Lança ValueError se inválido.
    Se device_id for informado, usa-o diretamente em vez de adivinhar via CSV."""
    if not channel:
        return  # Removing channel is always allowed
    tag_device = device_id or find_tag_device(tag)
    if not tag_device:
        return  # Tag not found in any device CSV — allow (backward compat)
    cfg = load_group_config()
    dev = cfg.get('devices', {}).get(tag_device, {})
    dev_channels = set(dev.get('channels', {}).keys())
    if channel not in dev_channels:
        raise ValueError(
            f"Canal '{channel}' nao pertence ao device '{tag_device}' da tag '{tag}'. "
            f"Canais disponiveis: {sorted(dev_channels)}"
        )


def patch_variable_override(tag: str, fields: dict, device_id: str | None = None) -> None:
    """
    Atualiza (ou cria) o override de uma variável individual.
    Campos suportados: enabled, channel.
    Campo delay_ms é ignorado silenciosamente.
    Valor None remove a chave correspondente do override.
    Override vazio após a operação → entrada removida completamente.
    Se device_id informado, opera no arquivo per-device.
    """
    allowed = {k: v for k, v in fields.items() if k != 'delay_ms'}
    if not allowed:
        return
    overrides = load_overrides(device_id)
    entry = overrides.setdefault(tag, {})
    for k, v in allowed.items():
        if v is None or v == '':
            entry.pop(k, None)   # None ou string vazia → remove a chave
        else:
            entry[k] = v
    if not entry:
        del overrides[tag]       # override vazio → remove a entrada
    save_overrides(overrides, device_id)
    logger.info("Override da tag '%s' atualizado: %s (device=%s)", tag, allowed, device_id)


# ── Variáveis: leitura mesclada ───────────────────────────────────────────────

def load_all_variables() -> list:
    """
    Retorna lista de todas as variáveis com configuração mesclada:
    CSV → variable_overrides (per-device com fallback para global).

    Canal efetivo = overrides[tag]['channel'].  None quando não atribuída.

    Cada item:
        tag, group, type, address, channel, history_size, enabled, source, device
    """
    cfg          = load_group_config()
    meta         = cfg.get('_meta', {})
    default_hist = meta.get('default_history_size', 100)
    global_overrides = load_overrides()       # global fallback
    channels_data = get_channels()
    devices      = cfg.get('devices', {})

    variables: list[dict] = []

    def _read_csv_files(csv_file_list: list, device_id: str | None, dev_overrides: dict, dev_channels: dict) -> None:
        for csv_name in csv_file_list:
            csv_path = os.path.join(_TABLES_DIR, csv_name)
            if not os.path.exists(csv_path):
                continue

            source = os.path.splitext(os.path.basename(csv_path))[0]

            df = pd.read_csv(csv_path, sep=',', dtype={'Modbus': str})
            df = df.dropna(subset=['Modbus', 'key', 'ObjecTag'])

            for _, row in df.iterrows():
                tag     = str(row['ObjecTag']).strip()
                group   = str(row['key']).strip()
                var_at  = str(row.get('At', '')).strip()
                modbus_raw = str(row['Modbus']).strip()

                # Parse endereço com possível sufixo de bit
                try:
                    address, bit_index = parse_modbus_address(modbus_raw)
                except (ValueError, TypeError):
                    address = None
                    bit_index = None

                classe  = str(row['Classe']).strip() if 'Classe' in row.index and pd.notna(row.get('Classe')) else None

                ov      = dev_overrides.get(tag, {})
                enabled = ov.get('enabled', True)
                channel = ov.get('channel')   # None quando não atribuída
                added   = ov.get('_added', False)

                # Use device channels for history_size, fallback to global channels_data
                hist = None
                if channel:
                    if channel in dev_channels:
                        hist = dev_channels[channel].get('history_size', default_hist)
                    else:
                        hist = channels_data.get(channel, {}).get('history_size', default_hist)

                variables.append({
                    'tag':          tag,
                    'group':        group,
                    'type':         var_at,
                    'address':      address,
                    'bit_index':    bit_index,
                    'channel':      channel,
                    'history_size': hist,
                    'enabled':      enabled,
                    'source':       source,
                    'device':       device_id,
                    'classe':       classe,
                    '_added':       added,
                })

    if devices:
        for dev_id, dev_cfg in devices.items():
            # Try per-device overrides first, fall back to global
            dev_overrides_path = _overrides_path(dev_id)
            if os.path.exists(dev_overrides_path):
                dev_overrides = load_overrides(dev_id)
            else:
                dev_overrides = global_overrides   # global fallback

            # Get channels from device, not global
            dev_channels = dev_cfg.get('channels', {})

            csv_files = dev_cfg.get('csv_files', [])
            _read_csv_files(csv_files, dev_id, dev_overrides, dev_channels)
    else:
        # backward compat: sem seção devices, usa os CSVs padrão
        _read_csv_files(['operacao.csv', 'configuracao.csv'], None, global_overrides, {})

    return variables


def load_device_variables(device_id: str, channel_filter: str | None = None) -> list[dict]:
    """
    Retorna variáveis de um device específico.
    Se channel_filter, filtra também por overrides com channel == channel_filter.
    Cada item: {tag, address, type, group, ...}
    """
    all_vars = load_all_variables()
    result = [v for v in all_vars if v.get('device') == device_id]
    if channel_filter:
        result = [v for v in result if v.get('channel') == channel_filter]
    return result


# ── Export / Import ───────────────────────────────────────────────────────────

_EXPORT_COLUMNS = ['tag', 'group', 'type', 'address', 'channel', 'history_size', 'enabled', 'source']
_EXPORT_HEADERS = ['Tag', 'Grupo', 'Tipo', 'Endereço', 'Canal', 'History size', 'Habilitado', 'Fonte']


def generate_export_xlsx() -> bytes:
    """Gera e retorna os bytes de um .xlsx com a configuração mesclada atual."""
    variables = load_all_variables()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Variáveis'

    ws.append(_EXPORT_HEADERS)

    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for var in variables:
        ws.append([var.get(col) for col in _EXPORT_COLUMNS])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_upload_xlsx(file_bytes: bytes) -> list:
    """
    Parseia um arquivo enviado pelo usuário — aceita .xlsx ou .csv.
    Detecta o formato pelos magic bytes (xlsx = ZIP → começa com b'PK').
    Colunas reconhecidas: Tag, Canal, History size, Habilitado (mínimo: Tag + Canal).
    Retorna lista de dicts com os campos reconhecidos.
    """
    _is_xlsx = file_bytes[:2] == b'PK'

    if _is_xlsx:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    else:
        import csv
        text = file_bytes.decode('utf-8-sig', errors='replace')
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=',;\t')
        except csv.Error:
            dialect = csv.excel   # fallback: vírgula padrão
        reader = csv.reader(io.StringIO(text), dialect)
        rows   = [tuple(row) for row in reader]

    if not rows:
        return []

    raw_headers = [str(h).strip() if h else '' for h in rows[0]]

    # Detecta formato: CSV nativo Modbus (tem 'ObjecTag') vs exportado pelo Hub (tem 'Tag'+'Canal')
    if 'ObjecTag' in raw_headers:
        header_map = {
            'ObjecTag': 'tag',
            'key':      'group',   # no CSV nativo, 'key' é o namespace/grupo
            'At':       'type',
            'Modbus':   'address',
        }
    else:
        header_map = {
            'Tag':          'tag',
            'Canal':        'channel',
            'History size': 'history_size',
            'Habilitado':   'enabled',
            'Grupo':        'group',
            'Fonte':        'source',
        }

    col_index: dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        if h in header_map:
            col_index[header_map[h]] = i

    if 'tag' not in col_index:
        raise ValueError(
            "Coluna de tag não encontrada. "
            "Esperado: 'Tag' (formato exportado) ou 'ObjecTag' (CSV Modbus nativo)."
        )

    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        item: dict = {}
        for field, idx in col_index.items():
            val = row[idx] if idx < len(row) else None
            if field == 'enabled':
                item[field] = bool(val) if val is not None else True
            elif field == 'history_size':
                item[field] = int(val) if val is not None else None
            else:
                item[field] = str(val).strip() if val is not None else None
        if item.get('tag'):
            result.append(item)

    return result


def apply_upload_config(rows: list) -> None:
    """
    Aplica configuração vinda de um upload (parse_upload_xlsx).

    Para cada linha: atualiza variable_overrides.json com channel e enabled.
    - channel vazio/None → remove a atribuição de canal
    - enabled=False → cria override; enabled=True → remove override de enabled
    - Override vazio após a operação → entrada removida
    """
    if not rows:
        return

    overrides = load_overrides()

    for row in rows:
        tag = row.get('tag', '').strip()
        if not tag:
            continue

        entry = overrides.get(tag, {})

        # Só atualiza o campo se a coluna estava presente no arquivo
        if 'channel' in row:
            channel = row['channel'] or None   # string vazia → None
            if channel:
                entry['channel'] = channel
            else:
                entry.pop('channel', None)

        if 'enabled' in row:
            if not row['enabled']:
                entry['enabled'] = False
            else:
                entry.pop('enabled', None)

        if entry:
            overrides[tag] = entry
        else:
            overrides.pop(tag, None)

    save_overrides(overrides)
    logger.info("apply_upload_config: %d linhas aplicadas.", len(rows))


# ── Edição de variável no CSV ─────────────────────────────────────────────

# Mapa: campo da API → coluna do CSV
_FIELD_TO_CSV_COL = {
    'tag':     'ObjecTag',
    'group':   'key',
    'type':    'At',
    'address': 'Modbus',
    'classe':  'Classe',
}


def update_csv_variable(tag: str, fields: dict, device_id: str | None = None) -> bool:
    """
    Atualiza campos de uma variável diretamente no CSV fonte.

    Localiza o CSV que contém a tag (restrito aos csv_files do device se informado),
    modifica a linha correspondente e reescreve o arquivo.

    Campos aceitos: tag, group, type, address, classe (mapeados para colunas CSV).
    Retorna True se a variável foi encontrada e atualizada, False caso contrário.
    """
    csv_fields = {_FIELD_TO_CSV_COL[k]: v for k, v in fields.items() if k in _FIELD_TO_CSV_COL}
    if not csv_fields:
        return False

    cfg = load_group_config()

    # Determina lista de CSVs a buscar
    if device_id:
        dev = cfg.get('devices', {}).get(device_id, {})
        csv_files = dev.get('csv_files', [])
    else:
        csv_files = []
        for dev_cfg in cfg.get('devices', {}).values():
            csv_files.extend(dev_cfg.get('csv_files', []))

    for csv_name in csv_files:
        csv_path = os.path.join(_TABLES_DIR, csv_name)
        if not os.path.exists(csv_path):
            continue
        try:
            df = pd.read_csv(csv_path, sep=',')
        except Exception:
            continue
        if 'ObjecTag' not in df.columns:
            continue

        mask = df['ObjecTag'].astype(str).str.strip() == tag
        if not mask.any():
            continue

        # Aplica alterações
        for col, val in csv_fields.items():
            if col in df.columns:
                df.loc[mask, col] = val
            elif col == 'Classe':
                # Classe pode não existir no CSV — adiciona coluna
                df[col] = ''
                df.loc[mask, col] = val

        # Se o tag foi renomeado, atualizar a coluna ObjecTag
        if 'ObjecTag' in csv_fields:
            df.loc[mask, 'ObjecTag'] = csv_fields['ObjecTag']

        df.to_csv(csv_path, index=False)
        logger.info("CSV '%s': tag '%s' atualizada — %s", csv_name, tag, csv_fields)
        return True

    return False


def add_csv_variable(device_id: str, csv_file: str, tag: str, group: str,
                     at_type: str, address: str, classe: str = '') -> dict:
    """
    Adiciona uma nova variável ao CSV de mapeamento Modbus.

    Valida que:
      - O device existe e csv_file pertence a ele
      - A tag não é duplicada no CSV
      - at_type é '%MB' ou '%MW'

    Determina Tipo (M/D) a partir de at_type:
      - '%MB' → 'M' (coil)
      - '%MW' → 'D' (register)

    Retorna dict da variável criada.
    """
    cfg = load_group_config()
    dev = cfg.get('devices', {}).get(device_id)
    if not dev:
        raise KeyError(f"Device '{device_id}' não encontrado.")
    if csv_file not in dev.get('csv_files', []):
        raise ValueError(f"CSV '{csv_file}' não pertence ao device '{device_id}'.")
    if at_type not in ('%MB', '%MW'):
        raise ValueError(f"Tipo '{at_type}' inválido. Use '%MB' ou '%MW'.")

    csv_path = os.path.join(_TABLES_DIR, csv_file)
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"Arquivo '{csv_file}' não encontrado em {_TABLES_DIR}.")

    df = pd.read_csv(csv_path, sep=',', dtype={'Modbus': str})

    # Verifica duplicata
    if 'ObjecTag' in df.columns:
        existing = df['ObjecTag'].astype(str).str.strip().values
        if tag in existing:
            raise ValueError(f"Tag '{tag}' já existe no CSV '{csv_file}'.")

    # Determina Tipo a partir de At
    tipo = 'M' if at_type == '%MB' else 'D'

    new_row = {
        'key': group,
        'ObjecTag': tag,
        'Tipo': tipo,
        'Modbus': str(address),
        'At': at_type,
    }
    if 'Classe' in df.columns or classe:
        new_row['Classe'] = classe or ''

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df.to_csv(csv_path, index=False)
    logger.info("CSV '%s': variável '%s' adicionada (group=%s, at=%s, addr=%s).",
                csv_file, tag, group, at_type, address)

    return {
        'tag': tag,
        'group': group,
        'type': at_type,
        'address': address,
        'classe': classe,
        'device': device_id,
        'csv_file': csv_file,
    }


def delete_csv_variable(tag: str, device_id: str | None = None) -> bool:
    """
    Remove uma variável do CSV fonte e seu override.

    Localiza o CSV que contém a tag (restrito aos csv_files do device se informado),
    remove a linha correspondente e reescreve o arquivo.
    Remove também a entrada do override per-device.

    Retorna True se a variável foi encontrada e removida, False caso contrário.
    """
    cfg = load_group_config()

    if device_id:
        dev = cfg.get('devices', {}).get(device_id, {})
        csv_files = dev.get('csv_files', [])
    else:
        csv_files = []
        for dev_cfg in cfg.get('devices', {}).values():
            csv_files.extend(dev_cfg.get('csv_files', []))

    for csv_name in csv_files:
        csv_path = os.path.join(_TABLES_DIR, csv_name)
        if not os.path.exists(csv_path):
            continue
        try:
            df = pd.read_csv(csv_path, sep=',', dtype={'Modbus': str})
        except Exception:
            continue
        if 'ObjecTag' not in df.columns:
            continue

        mask = df['ObjecTag'].astype(str).str.strip() == tag
        if not mask.any():
            continue

        df = df[~mask]
        df.to_csv(csv_path, index=False)

        # Remove override
        resolved_device = device_id or find_tag_device(tag)
        if resolved_device:
            overrides = load_overrides(resolved_device)
            if tag in overrides:
                del overrides[tag]
                save_overrides(overrides, resolved_device)

        logger.info("CSV '%s': variável '%s' removida.", csv_name, tag)
        return True

    return False


def rename_channel(old_name: str, new_name: str, device_id: str) -> None:
    """
    Renomeia um canal dentro de um device.

    Atualiza:
      1. group_config.json: devices[device_id].channels (renomeia a chave)
      2. variable_overrides_{device_id}.json: todas as tags com channel=old_name → new_name

    Valida que:
      - old_name existe nos channels do device
      - new_name não é vazio
      - Nem old_name nem new_name é canal de sistema
      - new_name não já existe nos channels do device
    """
    if old_name in SYSTEM_CHANNELS:
        raise ValueError(f"Canal '{old_name}' é um canal de sistema e não pode ser renomeado.")
    if new_name in SYSTEM_CHANNELS:
        raise ValueError(f"Canal '{new_name}' é um canal de sistema e não pode ser usado como destino.")
    if not new_name or not new_name.strip():
        raise ValueError("O novo nome do canal não pode ser vazio.")

    config = load_group_config()
    dev = config.get('devices', {}).get(device_id)
    if not dev:
        raise KeyError(f"Device '{device_id}' não encontrado.")

    channels = dev.get('channels', {})
    if old_name not in channels:
        raise KeyError(f"Canal '{old_name}' não encontrado no device '{device_id}'.")
    if new_name in channels:
        raise ValueError(f"Canal '{new_name}' já existe no device '{device_id}'.")

    # Renomeia a chave no channels do device
    channels[new_name] = channels.pop(old_name)
    save_group_config(config)

    # Atualiza overrides per-device
    overrides = load_overrides(device_id)
    changed = 0
    for tag, ov in overrides.items():
        if ov.get('channel') == old_name:
            ov['channel'] = new_name
            changed += 1
    if changed:
        save_overrides(overrides, device_id)

    logger.info("Canal '%s' renomeado para '%s' no device '%s' (%d overrides migrados).",
                old_name, new_name, device_id, changed)
